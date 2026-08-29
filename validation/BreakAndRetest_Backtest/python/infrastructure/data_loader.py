"""
FILE: infrastructure/data_loader.py
VERSION: 1.4
DATE: 2026-08-26
AUTHOR: Tony + Claude
LAYER: infrastructure
DESCRIPTION:
    Reads P_300 bulk VP export files (10_Pattern_<SYMBOL>.xlsx) read-only.
    Verified 22-column layout, 2-row header, newest-date-first row order
    (reversed here to chronological ascending for the strategy engine).

    Read-only against P_300's data/bulk/mine/ folder -- never writes back,
    never touches catalog.db or the P_300 pipeline.

CHANGELOG:
    - 2026-08-26 v1.4: Added _coerce_neural_index. VP writes 'n/a' (and
      possibly other non-up/down text) on export-time gaps, not always
      literally 'unknown'. Anything other than 'up'/'down' now normalizes
      to 'unknown', matching BulkBarRaw's own documented intent for that
      value. Observed on APP.
    - 2026-08-26 v1.3: Blank cells now coerce to 0.0 instead of raising --
      VP leaves numeric cells empty (not '0') during its pre-backfill
      period, same convention schemas_bulk.py's BulkBarRaw docstring
      already documents for pred_high/pred_low/triple-cross.
    - 2026-08-26 v1.2: Added _coerce_float to handle VP's literal '∞' /
      '-∞' string values (observed in roc_pct on early bars where the
      rate-of-change denominator is zero).
    - 2026-08-26 v1.1: Filename regex widened to allow underscores
      (10_Pattern_BRK_A.xlsx, 10_Pattern_BRK_B.xlsx were failing to match
      letters-only pattern).
    - 2026-08-26 v1.0: Initial build.
"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Literal

import openpyxl

from schemas import BulkBarRaw

logger = logging.getLogger(__name__)

_FILENAME_PATTERN = re.compile(r"10_Pattern_([A-Za-z_]+)\.xlsx$")
_EXPECTED_COLUMN_COUNT = 22
_DATA_START_ROW = 3  # row 1 = top headers, row 2 = sub headers, row 3 = first bar
_INF_STRINGS = {"∞", "inf", "infinity"}
_NEG_INF_STRINGS = {"-∞", "-inf", "-infinity"}


def extract_symbol(filepath: Path) -> str:
    """Pulls the ticker out of a 10_Pattern_<SYMBOL>.xlsx filename."""
    match = _FILENAME_PATTERN.search(filepath.name)
    if not match:
        raise ValueError(f"Filename does not match 10_Pattern_<SYMBOL>.xlsx: {filepath.name}")
    return match.group(1)


def _coerce_float(value, field_name: str, row_num: int) -> float:
    """Handles VP's literal infinity strings and blank pre-backfill cells.

    '∞' / '-∞' -- ratio's denominator was zero (seen on roc_pct).
    '' / None  -- VP's pre-backfill convention, mapped to 0.0 (same
        convention schemas_bulk.py documents for pred_high/pred_low/
        triple-cross columns).
    Anything else that still won't parse raises -- a genuine data
    problem, not a known VP export quirk.
    """
    if value is None:
        return 0.0
    if isinstance(value, str):
        stripped = value.strip().lower()
        if stripped == "":
            return 0.0
        if stripped in _INF_STRINGS:
            return float("inf")
        if stripped in _NEG_INF_STRINGS:
            return float("-inf")
        try:
            return float(stripped)
        except ValueError as exc:
            raise ValueError(
                f"row {row_num}: unparseable value {value!r} in {field_name}"
            ) from exc
    return float(value)


def _coerce_neural_index(value) -> Literal["up", "down", "unknown"]:
    """Normalizes VP's Neural Index direction text.

    Only 'up' and 'down' are real signals; every other export-time gap
    marker (blank, 'n/a', or anything else VP writes) collapses to
    'unknown' -- BulkBarRaw's documented catch-all for this field.
    """
    if isinstance(value, str) and value.strip().lower() in ("up", "down"):
        return value.strip().lower()  # type: ignore[return-value]
    return "unknown"


def load_bulk_file(filepath: Path) -> tuple[str, list[BulkBarRaw]]:
    """Reads one bulk VP export and returns (symbol, bars sorted ascending).

    Raises:
        ValueError: filename doesn't match convention, workbook fails to
            open, a row has the wrong column count, or a row fails
            BulkBarRaw validation (bad OHLC, unparseable numeric, etc.).
    """
    symbol = extract_symbol(filepath)
    logger.info("Loading %s from %s", symbol, filepath.name)

    try:
        workbook = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(min_row=_DATA_START_ROW, values_only=True))
    except Exception as exc:
        raise ValueError(f"Failed to open {filepath.name}: {exc}") from exc

    bars: list[BulkBarRaw] = []
    for row_num, row in enumerate(rows, start=_DATA_START_ROW):
        if row[0] is None:
            continue  # skip trailing blank rows
        if len(row) != _EXPECTED_COLUMN_COUNT:
            raise ValueError(
                f"{filepath.name} row {row_num}: expected "
                f"{_EXPECTED_COLUMN_COUNT} columns, got {len(row)}"
            )
        bars.append(_row_to_bar(row, row_num))

    bars.sort(key=lambda b: b.bar_date)  # file is newest-first; reverse to ascending
    logger.info("%s: %d bars, %s to %s", symbol, len(bars), bars[0].bar_date, bars[-1].bar_date)
    return symbol, bars


def _row_to_bar(row: tuple, row_num: int) -> BulkBarRaw:
    """Maps one raw xlsx row (22 columns, verified order) to BulkBarRaw."""
    return BulkBarRaw(
        bar_date=row[0].date(),
        stdiff=_coerce_float(row[1], "stdiff", row_num),
        mtdiff=_coerce_float(row[2], "mtdiff", row_num),
        ltdiff=_coerce_float(row[3], "ltdiff", row_num),
        open=_coerce_float(row[4], "open", row_num),
        high=_coerce_float(row[5], "high", row_num),
        low=_coerce_float(row[6], "low", row_num),
        close=_coerce_float(row[7], "close", row_num),
        pred_high=_coerce_float(row[8], "pred_high", row_num),
        pred_low=_coerce_float(row[9], "pred_low", row_num),
        volume=_coerce_float(row[10], "volume", row_num),
        williams_emai=_coerce_float(row[11], "williams_emai", row_num),
        psi=_coerce_float(row[12], "psi", row_num),
        roc_pct=_coerce_float(row[13], "roc_pct", row_num),
        neural_index=_coerce_neural_index(row[14]),
        neural_x_max=_coerce_float(row[15], "neural_x_max", row_num),
        tc_short=_coerce_float(row[16], "tc_short", row_num),
        tc_medium=_coerce_float(row[17], "tc_medium", row_num),
        tc_long=_coerce_float(row[18], "tc_long", row_num),
        pred_high_diff=_coerce_float(row[19], "pred_high_diff", row_num),
        pred_low_diff=_coerce_float(row[20], "pred_low_diff", row_num),
        pred_range=_coerce_float(row[21], "pred_range", row_num),
    )


def find_bulk_files(data_dir: Path, glob_pattern: str) -> list[Path]:
    """Lists all bulk export files in data_dir matching glob_pattern."""
    files = sorted(data_dir.glob(glob_pattern))
    logger.info("Found %d files matching %s in %s", len(files), glob_pattern, data_dir)
    return files
