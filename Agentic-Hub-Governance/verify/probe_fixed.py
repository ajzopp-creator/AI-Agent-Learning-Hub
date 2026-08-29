import openpyxl
wb = openpyxl.load_workbook(r"C:\Users\Trader\AI-Agent-Learning-Hub\projects\P_115_BuytheDipTradingSystem\data\P_115_TrackerDashboard_V3.xlsx", data_only=True)
ws = wb["Tracker Log"]
header = [c.value for c in ws[1]]
idx = {}
for i, h in enumerate(header):
    if h:
        idx[h] = i
rows = list(ws.iter_rows(min_row=2, values_only=True))
lines = []
lines.append("Total data rows: " + str(len(rows)))
for r in rows[-25:]:
    d = r[idx["Date"]]
    sym = r[idx["Symbol"]]
    src = r[idx["SignalSource"]]
    verdict = r[idx["Step1Verdict"]]
    lines.append("Date=" + str(d) + " Symbol=" + str(sym) + " Source=" + str(src) + " Verdict=" + str(verdict))
outp = r"C:\Users\Trader\AI-Agent-Learning-Hub\Agentic-Hub-Governance\verify\diag_output.txt"
with open(outp, "w", encoding="utf-8") as f:
    f.write(chr(10).join(lines))
print("done")
