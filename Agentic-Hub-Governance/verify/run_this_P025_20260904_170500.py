"""
PEH run_this — P_025 WO-P025-EN.001
Timestamp: 20260904_170500

Does NOT copy or rmtree python\\. Assumes code already extracted
onto the P_025 project. Runs yearly build + format_analytics + checks.
"""

from __future__ import annotations

import os
import sys
from datetime import datetime
from pathlib import Path

HUB_ROOT = Path(os.environ.get("HUB_ROOT", r"C:\Users\Trader\AI-Agent-Learning-Hub"))
PROJECT_ROOT = HUB_ROOT / "projects" / "P_025_AJZ_Institutional_Portfolio_Tracker"
PYTHON_ROOT = PROJECT_ROOT / "python"
OUTPUT_DIR = PROJECT_ROOT / "output"
WORKBOOK_PATH = OUTPUT_DIR / "P_025_Portfolio_BUILT.xlsx"

EXPECTED_LAKE = [
    "Trade_Log",
    "Market_Data",
    "Reference_Data",
    "Daily_Units",
    "Daily_Cash",
    "Daily_Invested",
    "Cost_Basis",
    "Fifo_Lots",
    "Fifo_Cost",
]


def write_done(status: str, exit_code: int) -> None:
    done_path = Path(__file__).with_suffix(Path(__file__).suffix + ".done")
    done_path.write_text(
        f"status={status}\nexit_code={exit_code}\ntimestamp={datetime.now().isoformat()}\n",
        encoding="utf-8",
    )


def fail(msg: str) -> None:
    print(f"FAIL: {msg}")
    write_done("FAIL", 1)
    sys.exit(1)


def main() -> int:
    print("=" * 60)
    print("P_025 PEH 20260904_170500  WO-P025-EN.001")
    print("=" * 60)
    print(f"HUB_ROOT    : {HUB_ROOT}")
    print(f"Python root : {PYTHON_ROOT}")
    print(f"Workbook    : {WORKBOOK_PATH}")

    if "OneDrive" in str(HUB_ROOT) and not os.environ.get("HUB_ROOT"):
        fail("HUB_ROOT resolved to OneDrive without an explicit env override")

    if not PYTHON_ROOT.exists():
        fail(f"python\\ folder missing: {PYTHON_ROOT}")

    if str(PYTHON_ROOT) not in sys.path:
        sys.path.insert(0, str(PYTHON_ROOT))

    import config as cfg

    if "OneDrive" in cfg._DEFAULT_HUB_ROOT:
        fail("config._DEFAULT_HUB_ROOT still contains OneDrive")
    print(f"  OK  default hub = {cfg._DEFAULT_HUB_ROOT}")
    print(f"  OK  IRA_FEED_READY = {cfg.IRA_FEED_READY}")

    print("\n=== pytest ===")
    import subprocess

    rc = subprocess.run(
        [sys.executable, "-m", "pytest", "-q", str(PYTHON_ROOT / "tests")],
        check=False,
    ).returncode
    if rc != 0:
        fail(f"pytest failed with code {rc}")
    print("  OK  pytest")

    print("\n=== yearly build ===")
    try:
        from application.build_portfolio import run_full_build

        snapshot = run_full_build(mode="yearly")
        print(f"  OK  trades={len(snapshot.trades)}")
        print(f"  OK  market={len(snapshot.market_data)}")
        print(f"  OK  fifo_lots={len(snapshot.fifo_lots)}")
        print(f"  OK  fifo_cost={len(snapshot.fifo_cost)}")
        print(f"  OK  cost_basis={len(snapshot.cost_basis)}")
    except Exception as exc:
        fail(f"Build raised {type(exc).__name__}: {exc}")

    if not WORKBOOK_PATH.exists():
        fail(f"Workbook missing: {WORKBOOK_PATH}")
    print(f"  OK  workbook {WORKBOOK_PATH.stat().st_size / 1024:.1f} KB")

    print("\n=== format analytics ===")
    try:
        from application.format_analytics import run_format_analytics

        dest = run_format_analytics()
        print(f"  OK  analytics → {dest}")
    except Exception as exc:
        fail(f"format_analytics raised {type(exc).__name__}: {exc}")

    print("\n=== sheet checks ===")
    from openpyxl import load_workbook

    wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=False)
    names = set(wb.sheetnames)
    missing = [s for s in EXPECTED_LAKE if s not in names]
    if missing:
        fail(f"Missing Data Lake sheets: {missing}")
    print(f"  OK  lake sheets {EXPECTED_LAKE}")
    tl = wb["Trade_Log"]
    if tl.max_row < 2:
        fail("Trade_Log has no data rows")
    print(f"  OK  Trade_Log max_row={tl.max_row}")
    wb.close()

    wb2 = load_workbook(dest, read_only=True, data_only=False)
    corr = wb2["Correlation"]
    sample = corr.cell(5, 3).value
    if not (isinstance(sample, str) and "CORREL" in sample):
        fail(f"Correlation off-diagonal is not CORREL: {sample!r}")
    print("  OK  Correlation has CORREL formulas")
    geo = wb2["Geographic_Exposure"].cell(1, 1).value or ""
    if "placeholder" in str(geo).lower():
        fail("Geographic still placeholder")
    print(f"  OK  Geographic title={geo}")
    stress = wb2["Stress_Testing"].cell(1, 1).value or ""
    if "placeholder" in str(stress).lower():
        fail("Stress still placeholder")
    print(f"  OK  Stress title={stress}")
    pos_cost = wb2["Positions"].cell(3, 5).value or ""
    if "Fifo_Cost" not in str(pos_cost):
        fail(f"Positions cost is not Fifo_Cost SUMIF: {pos_cost!r}")
    print("  OK  Positions cost → Fifo_Cost")
    wb2.close()

    print("\n" + "=" * 60)
    print("PASS")
    print("=" * 60)
    write_done("PASS", 0)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
