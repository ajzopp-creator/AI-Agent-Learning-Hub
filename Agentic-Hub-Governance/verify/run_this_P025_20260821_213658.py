"""
PEH run_this — P_025 Full Build Test
Timestamp: 20260821_213658

Runs the full portfolio build against the live P_020 database and
verifies the output workbook.

Prints PASS only when the workbook exists and contains the expected
Data Lake sheets with data.
Writes a sibling .done marker on completion.
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
HUB_ROOT = Path(r"C:\Users\Trader\AI-Agent-Learning-Hub")
PROJECT_ROOT = HUB_ROOT / "projects" / "P_025_AJZ_Institutional_Portfolio_Tracker"
PYTHON_ROOT = PROJECT_ROOT / "python"
OUTPUT_DIR = PROJECT_ROOT / "output"
WORKBOOK_PATH = OUTPUT_DIR / "P_025_Portfolio_BUILT.xlsx"

EXPECTED_DATA_LAKE = [
    "Trade_Log",
    "Market_Data",
    "Reference_Data",
    "Daily_Units",
    "Daily_Cash",
]

EXPECTED_ANALYTICS = [
    "Dashboard",
    "Positions",
    "Equity_Curve",
    "Sector_Exposure",
    "Geographic_Exposure",
    "Correlation",
    "Risk_Metrics",
    "Stress_Testing",
    "Investment_Theses",
]


def write_done(status: str, exit_code: int) -> None:
    done_path = Path(__file__).with_suffix(Path(__file__).suffix + ".done")
    done_path.write_text(
        f"status={status}\nexit_code={exit_code}\ntimestamp={datetime.now().isoformat()}\n",
        encoding="utf-8",
    )


def fail(msg: str) -> None:
    print(f"FAIL: {msg}")
    write_done("FAIL", 1)
    sys.exit(1)


def main() -> int:
    print("=" * 60)
    print("P_025 Full Build Test — 20260821_213658")
    print("=" * 60)
    print(f"Python root : {PYTHON_ROOT}")
    print(f"Workbook    : {WORKBOOK_PATH}")

    if not PYTHON_ROOT.exists():
        fail(f"python\\ folder missing: {PYTHON_ROOT}")

    # Make the package importable
    if str(PYTHON_ROOT) not in sys.path:
        sys.path.insert(0, str(PYTHON_ROOT))

    # ------------------------------------------------------------------
    # 1. Run the full build
    # ------------------------------------------------------------------
    print("\n=== Running full build ===")
    try:
        from application.build_portfolio import run_full_build
        snapshot = run_full_build()
        print(f"  OK  Build returned. Trades in snapshot: {len(snapshot.trades)}")
        print(f"  OK  Market data rows : {len(snapshot.market_data)}")
        print(f"  OK  Reference rows   : {len(snapshot.reference_data)}")
        print(f"  OK  Daily units rows : {len(snapshot.daily_units)}")
        print(f"  OK  Daily cash rows  : {len(snapshot.daily_cash)}")
    except Exception as exc:
        fail(f"Build raised exception: {type(exc).__name__}: {exc}")

    # ------------------------------------------------------------------
    # 2. Verify workbook exists
    # ------------------------------------------------------------------
    print("\n=== Verifying workbook ===")
    if not WORKBOOK_PATH.exists():
        fail(f"Workbook was not created: {WORKBOOK_PATH}")
    size_kb = WORKBOOK_PATH.stat().st_size / 1024
    print(f"  OK  Workbook exists ({size_kb:.1f} KB)")

    # ------------------------------------------------------------------
    # 3. Check sheets
    # ------------------------------------------------------------------
    try:
        from openpyxl import load_workbook
        wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=False)
        sheetnames = set(wb.sheetnames)
        print(f"  OK  Sheets present: {sorted(sheetnames)}")

        missing_lake = [s for s in EXPECTED_DATA_LAKE if s not in sheetnames]
        if missing_lake:
            fail(f"Missing Data Lake sheets: {missing_lake}")
        print("  OK  All five Data Lake sheets present")

        missing_analytics = [s for s in EXPECTED_ANALYTICS if s not in sheetnames]
        if missing_analytics:
            print(f"  WARN Missing Analytics placeholders: {missing_analytics}")
        else:
            print("  OK  All Analytics placeholder sheets present")

        # Quick row check on Trade_Log
        ws = wb["Trade_Log"]
        row_count = ws.max_row
        print(f"  OK  Trade_Log max_row = {row_count}")
        if row_count < 2:
            fail("Trade_Log has no data rows")
        wb.close()
    except Exception as exc:
        fail(f"Workbook inspection failed: {type(exc).__name__}: {exc}")

    print("\n" + "=" * 60)
    print("PASS")
    print("=" * 60)
    write_done("PASS", 0)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
