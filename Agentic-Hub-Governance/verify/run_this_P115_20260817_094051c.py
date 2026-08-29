import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\Trader\AI-Agent-Learning-Hub\projects\P_115_BuytheDipTradingSystem\data\P_115_TrackerDashboard_V3.xlsx', data_only=True)
ws = wb['Tracker Log']
header = [c.value for c in ws[1]]
idx = {h: i for i, h in enumerate(header) if h}
rows = list(ws.iter_rows(min_row=2, values_only=True))
last20 = rows[-20:]
with open(r'C:\Users\Trader\AI-Agent-Learning-Hub\Agentic-Hub-Governance\verify\last_rows_20260817_094051.txt', 'w', encoding='utf-8') as f:
    f.write(f'Total data rows: {len(rows)}\n')
    for r in last20:
        f.write(f\"Date={r[idx['Date']]!r} Symbol={r[idx['Symbol']]!r} Source={r[idx['SignalSource']]!r} Verdict={r[idx['Step1Verdict']]!r}\n\")
print('done')
