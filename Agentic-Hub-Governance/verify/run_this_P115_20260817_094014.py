import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook(r'C:\Users\Trader\AI-Agent-Learning-Hub\projects\P_115_BuytheDipTradingSystem\data\P_115_TrackerDashboard_V3.xlsx', data_only=True)
out = []
for sheet in wb.sheetnames:
    ws = wb[sheet]
    header = [c.value for c in ws[1]]
    out.append(f'SHEET: {sheet} | rows={ws.max_row} | cols={ws.max_column}')
    out.append('HEADER: ' + ' | '.join(str(h) for h in header))

with open(r'C:\Users\Trader\AI-Agent-Learning-Hub\Agentic-Hub-Governance\verify\tracker_probe_20260817_094014.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(out))
print('done')
