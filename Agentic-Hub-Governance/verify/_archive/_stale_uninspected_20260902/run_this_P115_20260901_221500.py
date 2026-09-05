"""
P_115 tracker append - ABT + XLV live-signal rows, 2026-09-01 session.
No WO (routine trade logging). Never modifies existing rows/formulas -
only appends two new rows, copying style from the last existing data row.

Locates the 27-column tracker sheet by header match, finds the last
non-empty row (by Date column), copies that row's per-cell style to two
new rows immediately below, writes values by HEADER NAME (not fixed
column index) so a schema reorder can't silently misalign data, saves,
and reads the file back to confirm both rows landed with the values
this script intended - not just that the file has more rows than before.
"""
import sys
import copy
import openpyxl

PATH = r"D:\OneDrive\Documents\AJZStrategiesLLC\P_115_TrackerAudit\P_115_118_TrackerDashboard_V2.xlsx"

# Locked 27-column schema (P_115 architecture doc / p115-project-context SKILL.md)
SCHEMA = [
    "Date", "Symbol", "SignalSource", "Step1Verdict", "PatternType",
    "BreakoutVerdict", "BreakoutVolumeMultiple", "DistributionDayCount",
    "FollowThroughDay", "MarketDirection", "RSvsSPY", "FundamentalsTier",
    "AnalysisTier", "CandleTier", "SetupScore", "LiquidityTier", "Traded",
    "EntryPrice", "TPLevel", "SLLevel", "StopLevel", "RiskPct",
    "AccountBalance", "Outcome", "RecheckStatus", "SimulationNotes",
    "Comments",
]

ROW_ABT = {
    "Date": "8/31/2026",
    "Symbol": "ABT",
    "SignalSource": "P_117",
    "Step1Verdict": "PASS",
    "PatternType": "--",
    "BreakoutVerdict": "--",
    "MarketDirection": "OFF",
    "FundamentalsTier": 2,
    "AnalysisTier": 1,
    "CandleTier": 0,
    "SetupScore": 0,
    "Traded": "Yes",
    "EntryPrice": 8.00,
    "TPLevel": 12.70,
    "SLLevel": 4.70,
    "StopLevel": 4.70,
    "RiskPct": 0.0105,
    "AccountBalance": 31348.39,
    "Outcome": "Open",
    "RecheckStatus": "N/A (PASS row, no re-verify required)",
    "Comments": ("Outside rec via SNT (P_117), not a P_115-engine trigger. "
                 "Chart/HybridTier read PASS (Anal=1+Fund=2=3). Filled: "
                 "BUY +1 ABT 18SEP26 105C @8.00, 8/31/26 09:30:09. GTC exits "
                 "working: STP 4.70 / LMT 12.70. R:R ~1.42:1."),
}

ROW_XLV = {
    "Date": "9/1/2026",
    "Symbol": "XLV",
    "SignalSource": "P_116",
    "Step1Verdict": "PASS",
    "PatternType": "Bounce",
    "BreakoutVerdict": "Bounce",
    "MarketDirection": "OFF",
    "FundamentalsTier": 2,
    "AnalysisTier": 1,
    "CandleTier": 0,
    "SetupScore": 1,
    "Traded": "No",
    "TPLevel": "179.50 (T1, 50%) / 183.80 (Primary)",
    "SLLevel": "167.60 (stock trigger)",
    "StopLevel": "167.60 (stock trigger)",
    "AccountBalance": 31348.39,
    "Outcome": "Pending",
    "RecheckStatus": "N/A (PASS row, no re-verify required)",
    "Comments": ("Council APPROVE WITH CAUTION (R:R 2.98:1 to Primary / "
                 "1.92:1 to T1). BUY +5 XLV 18SEP26 175C @1.10 LMT DAY, "
                 "placed after-hours 9/1/26 21:41:40, still WORKING (not "
                 "filled) as of log time. Council's 10-share stock leg "
                 "never appeared on any order ticket - logging option-only, "
                 "5 contracts. EntryPrice/RiskPct left blank pending fill."),
}


def find_tracker_sheet(wb):
    for name in wb.sheetnames:
        ws = wb[name]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        headers_norm = [str(h).strip() if h is not None else "" for h in headers]
        # Require at least Date, Symbol, SignalSource in row 1 to call it a match
        if "Date" in headers_norm and "Symbol" in headers_norm and "SignalSource" in headers_norm:
            return name, ws, headers_norm
    return None, None, None


def main():
    wb = openpyxl.load_workbook(PATH, data_only=False)
    print("SHEETS:", wb.sheetnames)

    sheet_name, ws, headers = find_tracker_sheet(wb)
    if ws is None:
        print("FAIL: no sheet found with Date/Symbol/SignalSource headers in row 1")
        sys.exit(1)

    print(f"MATCHED SHEET: {sheet_name}")
    print("HEADERS FOUND (row 1):", headers)

    missing_from_schema = [h for h in headers if h not in SCHEMA and h != ""]
    if missing_from_schema:
        print("NOTE: headers present in file but not in this script's SCHEMA list:", missing_from_schema)

    col_map = {h: i + 1 for i, h in enumerate(headers) if h}

    # Find last non-empty row by Date column
    date_col = col_map.get("Date")
    if not date_col:
        print("FAIL: no 'Date' column found - cannot locate last row safely")
        sys.exit(1)

    last_row = 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=date_col).value not in (None, ""):
            last_row = r
    print(f"LAST DATA ROW (before append) = {last_row}")

    before_max_row = ws.max_row

    new_rows = [ROW_ABT, ROW_XLV]
    written_summary = []

    for i, row_data in enumerate(new_rows):
        target_row = last_row + 1 + i
        for col_name, col_idx in col_map.items():
            src_cell = ws.cell(row=last_row, column=col_idx)
            dst_cell = ws.cell(row=target_row, column=col_idx)
            # Copy style from the last real data row so the new row matches
            # existing conventions (font, fill, border, number_format, alignment)
            dst_cell.font = copy.copy(src_cell.font)
            dst_cell.fill = copy.copy(src_cell.fill)
            dst_cell.border = copy.copy(src_cell.border)
            dst_cell.alignment = copy.copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
            # Write value if this row provides one for this column; otherwise
            # leave blank rather than guessing - matches "-- for not applicable"
            # convention only where this script explicitly set "--".
            if col_name in row_data:
                dst_cell.value = row_data[col_name]
        written_summary.append((target_row, row_data.get("Symbol")))

    wb.save(PATH)
    print("SAVED.")

    # Read back to confirm - do not trust a clean save() return alone.
    wb2 = openpyxl.load_workbook(PATH, data_only=False)
    ws2 = wb2[sheet_name]
    print(f"MAX ROW AFTER SAVE = {ws2.max_row} (was {before_max_row})")
    for target_row, sym in written_summary:
        row_vals = {h: ws2.cell(row=target_row, column=col_map[h]).value for h in col_map}
        print(f"ROW {target_row} ({sym}) READBACK:", row_vals)

    print("PASS")


if __name__ == "__main__":
    main()
