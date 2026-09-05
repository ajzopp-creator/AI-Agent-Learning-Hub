"""
PEH run_this — P_025 Analytics Layer (versioned workbook)
Timestamp: 20260821_222625

Applies Bloomberg-style formatting + core Analytics formulas to the
existing Data Lake workbook and saves a versioned copy.

Prints PASS when the versioned workbook exists and key sheets are populated.
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

HUB_ROOT = Path(r"C:\Users\Trader\AI-Agent-Learning-Hub")
PROJECT_ROOT = HUB_ROOT / "projects" / "P_025_AJZ_Institutional_Portfolio_Tracker"
PYTHON_ROOT = PROJECT_ROOT / "python"
SOURCE_WB = PROJECT_ROOT / "output" / "P_025_Portfolio_BUILT.xlsx"


def write_done(status: str, exit_code: int) -> None:
    done_path = Path(__file__).with_suffix(Path(__file__).suffix + ".done")
    done_path.write_text(
        f"status={status}\nexit_code={exit_code}\ntimestamp={datetime.now().isoformat()}\n",
        encoding="utf-8",
    )


def fail(msg: str) -> None:
    print(f"FAIL: {msg}")
    write_done("FAIL", 1)
    sys.exit(1)


def main() -> int:
    print("=" * 60)
    print("P_025 Analytics Layer — 20260821_222625")
    print("=" * 60)

    if not PYTHON_ROOT.exists():
        fail(f"python\\ folder missing: {PYTHON_ROOT}")
    if not SOURCE_WB.exists():
        fail(f"Source workbook missing: {SOURCE_WB}")

    if str(PYTHON_ROOT) not in sys.path:
        sys.path.insert(0, str(PYTHON_ROOT))

    # Verify new modules are present
    for rel in (
        "domain/formula_templates.py",
        "infrastructure/excel_formatter.py",
        "application/format_analytics.py",
    ):
        if not (PYTHON_ROOT / rel).exists():
            fail(
                f"Missing {rel}. Extract the updated P_025_python_source.zip "
                "over the python\\ folder and re-run."
            )
    print("  OK  Analytics modules present")

    print("\n=== Running format_analytics ===")
    try:
        from application.format_analytics import run_format_analytics
        dest = run_format_analytics(source_path=SOURCE_WB)
        print(f"  OK  Versioned workbook written → {dest}")
    except Exception as exc:
        fail(f"format_analytics raised: {type(exc).__name__}: {exc}")

    if not dest.exists():
        fail(f"Expected output not found: {dest}")

    size_mb = dest.stat().st_size / (1024 * 1024)
    print(f"  OK  Size {size_mb:.2f} MB")

    # Quick structural check
    try:
        from openpyxl import load_workbook
        wb = load_workbook(dest, read_only=True, data_only=False)
        for required in ("Positions", "Equity_Curve", "Dashboard", "Risk_Metrics"):
            if required not in wb.sheetnames:
                fail(f"Missing sheet after format: {required}")
            ws = wb[required]
            if ws.max_row < 2:
                fail(f"{required} appears empty (max_row={ws.max_row})")
            print(f"  OK  {required}: max_row={ws.max_row}, max_col={ws.max_column}")
        wb.close()
    except Exception as exc:
        fail(f"Inspection failed: {type(exc).__name__}: {exc}")

    print("\n" + "=" * 60)
    print("PASS")
    print("=" * 60)
    write_done("PASS", 0)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
