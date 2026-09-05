"""
P_115 tracker inspect - read-only, no writes to production file.
Purpose: identify sheet name, header row, last data row, and formatting
of an existing row so the ABT/XLV append matches conventions exactly.
"""
import openpyxl

PATH = r"D:\OneDrive\Documents\AJZStrategiesLLC\P_115_TrackerAudit\P_115_118_TrackerDashboard_V2.xlsx"

wb = openpyxl.load_workbook(PATH, data_only=False)
print("SHEETS:", wb.sheetnames)

for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n--- {name} --- dims={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column}")

# Assume first sheet is the tracker unless a name obviously matches
ws = wb[wb.sheetnames[0]]
print("\nHEADER ROW 1:")
for c in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=c)
    print(c, repr(cell.value))

last_row = ws.max_row
print(f"\nLAST ROW = {last_row}")
print("LAST ROW VALUES:")
for c in range(1, ws.max_column + 1):
    cell = ws.cell(row=last_row, column=c)
    print(c, repr(cell.value), "font=", cell.font.name, cell.font.size, "fill=", cell.fill.fgColor.rgb if cell.fill else None, "numfmt=", cell.number_format)

print("\nSECOND-TO-LAST ROW VALUES (for pattern comparison):")
for c in range(1, ws.max_column + 1):
    cell = ws.cell(row=last_row-1, column=c)
    print(c, repr(cell.value))

print("PASS")
