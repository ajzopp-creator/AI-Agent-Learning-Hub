"""
PEH run_this — P_025 Analytics split verification
Timestamp: 20260821_223516

1. Confirms the split modules exist and are under 300 lines.
2. Re-runs format_analytics to produce a fresh versioned workbook.
3. Verifies key Analytics sheets are populated.

Prints PASS only when all checks succeed.
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

HUB_ROOT = Path(r"C:\Users\Trader\AI-Agent-Learning-Hub")
PROJECT_ROOT = HUB_ROOT / "projects" / "P_025_AJZ_Institutional_Portfolio_Tracker"
PYTHON_ROOT = PROJECT_ROOT / "python"
SOURCE_WB = PROJECT_ROOT / "output" / "P_025_Portfolio_BUILT.xlsx"

MAX_LINES = 300
REQUIRED = [
    "application/format_analytics.py",
    "infrastructure/analytics_sheets.py",
    "infrastructure/excel_formatter.py",
    "domain/formula_templates.py",
]


def write_done(status: str, exit_code: int) -> None:
    done_path = Path(__file__).with_suffix(Path(__file__).suffix + ".done")
    done_path.write_text(
        f"status={status}\nexit_code={exit_code}\ntimestamp={datetime.now().isoformat()}\n",
        encoding="utf-8",
    )


def fail(msg: str) -> None:
    print(f"FAIL: {msg}")
    write_done("FAIL", 1)
    sys.exit(1)


def main() -> int:
    print("=" * 60)
    print("P_025 Analytics Split Verification — 20260821_223516")
    print("=" * 60)

    if not PYTHON_ROOT.exists():
        fail(f"python\\ missing: {PYTHON_ROOT}")
    if not SOURCE_WB.exists():
        fail(f"Source workbook missing: {SOURCE_WB}")

    # 1. Module presence + line-count gate
    print("\n=== Module line counts ===")
    for rel in REQUIRED:
        path = PYTHON_ROOT / rel
        if not path.exists():
            fail(f"Missing {rel}. Extract updated zip over python\\ and re-run.")
        lines = len(path.read_text(encoding="utf-8").splitlines())
        status = "OK" if lines <= MAX_LINES else "FAIL"
        print(f"  {status}  {rel}: {lines} lines")
        if lines > MAX_LINES:
            fail(f"{rel} exceeds {MAX_LINES}-line hard limit ({lines})")

    if str(PYTHON_ROOT) not in sys.path:
        sys.path.insert(0, str(PYTHON_ROOT))

    # 2. Run formatter
    print("\n=== Running format_analytics ===")
    try:
        from application.format_analytics import run_format_analytics
        dest = run_format_analytics(source_path=SOURCE_WB)
        print(f"  OK  Versioned workbook → {dest}")
    except Exception as exc:
        fail(f"format_analytics raised: {type(exc).__name__}: {exc}")

    if not dest.exists():
        fail(f"Output not found: {dest}")
    size_mb = dest.stat().st_size / (1024 * 1024)
    print(f"  OK  Size {size_mb:.2f} MB")

    # 3. Sheet checks
    print("\n=== Sheet checks ===")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(dest, read_only=True, data_only=False)
        for name in ("Positions", "Equity_Curve", "Dashboard", "Risk_Metrics"):
            if name not in wb.sheetnames:
                fail(f"Missing sheet: {name}")
            ws = wb[name]
            if ws.max_row < 2:
                fail(f"{name} empty (max_row={ws.max_row})")
            print(f"  OK  {name}: max_row={ws.max_row}, max_col={ws.max_column}")
        wb.close()
    except Exception as exc:
        fail(f"Inspection failed: {type(exc).__name__}: {exc}")

    print("\n" + "=" * 60)
    print("PASS")
    print("=" * 60)
    write_done("PASS", 0)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
