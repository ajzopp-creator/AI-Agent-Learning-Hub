"""
PEH run_this — P_025 Workbook Inspection
Timestamp: 20260821_220300

Read-only inspection of the live P_025_Portfolio_BUILT.xlsx.
Prints sheet inventory, row counts, column headers, and sample data.
Never modifies the workbook.
Prints PASS at the end.
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

HUB_ROOT = Path(r"C:\Users\Trader\AI-Agent-Learning-Hub")
WORKBOOK_PATH = (
    HUB_ROOT
    / "projects"
    / "P_025_AJZ_Institutional_Portfolio_Tracker"
    / "output"
    / "P_025_Portfolio_BUILT.xlsx"
)


def write_done(status: str, exit_code: int) -> None:
    done_path = Path(__file__).with_suffix(Path(__file__).suffix + ".done")
    done_path.write_text(
        f"status={status}\nexit_code={exit_code}\ntimestamp={datetime.now().isoformat()}\n",
        encoding="utf-8",
    )


def fail(msg: str) -> None:
    print(f"FAIL: {msg}")
    write_done("FAIL", 1)
    sys.exit(1)


def main() -> int:
    print("=" * 70)
    print("P_025 Workbook Inspection — 20260821_220300")
    print("=" * 70)
    print(f"Workbook: {WORKBOOK_PATH}")

    if not WORKBOOK_PATH.exists():
        fail(f"Workbook not found: {WORKBOOK_PATH}")

    size_mb = WORKBOOK_PATH.stat().st_size / (1024 * 1024)
    print(f"Size    : {size_mb:.2f} MB\n")

    try:
        from openpyxl import load_workbook
        wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=False)
    except Exception as exc:
        fail(f"Could not open workbook: {type(exc).__name__}: {exc}")

    print("--- Sheet Inventory ---")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  {name:25s}  max_row={ws.max_row:6d}  max_col={ws.max_column:3d}")
    print()

    # Detailed look at each Data Lake sheet
    data_lake = ["Trade_Log", "Market_Data", "Reference_Data", "Daily_Units", "Daily_Cash"]

    for sheet_name in data_lake:
        if sheet_name not in wb.sheetnames:
            print(f"WARNING: {sheet_name} missing")
            continue
        ws = wb[sheet_name]
        print(f"=== {sheet_name} ===")
        # Header row
        headers = []
        for col in range(1, min(ws.max_column + 1, 25)):
            val = ws.cell(1, col).value
            headers.append(str(val) if val is not None else "")
        print(f"  Headers ({len(headers)}): {headers}")
        # First 3 data rows
        for row in range(2, min(ws.max_row + 1, 5)):
            values = []
            for col in range(1, min(ws.max_column + 1, 12)):
                val = ws.cell(row, col).value
                values.append(str(val) if val is not None else "")
            print(f"  Row {row}: {values}")
        if ws.max_row > 4:
            print(f"  ... ({ws.max_row - 1} data rows total)")
        print()

    # Quick Analytics check
    print("--- Analytics placeholders (first cell of each) ---")
    analytics = [
        "Dashboard", "Positions", "Equity_Curve", "Sector_Exposure",
        "Geographic_Exposure", "Correlation", "Risk_Metrics",
        "Stress_Testing", "Investment_Theses",
    ]
    for name in analytics:
        if name in wb.sheetnames:
            val = wb[name].cell(1, 1).value
            print(f"  {name:25s}  A1={val!r}")
        else:
            print(f"  {name:25s}  MISSING")

    wb.close()

    print("\n" + "=" * 70)
    print("PASS")
    print("=" * 70)
    write_done("PASS", 0)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
