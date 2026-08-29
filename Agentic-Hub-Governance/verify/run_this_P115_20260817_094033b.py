import openpyxl
from datetime import datetime, date

wb = openpyxl.load_workbook(r'C:\Users\Trader\AI-Agent-Learning-Hub\projects\P_115_BuytheDipTradingSystem\data\P_115_TrackerDashboard_V3.xlsx', data_only=True)
ws = wb['Tracker Log']
header = [c.value for c in ws[1]]
idx = {h: i for i, h in enumerate(header) if h}

cutoff = date(2026, 7, 24)
rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    d = r[idx['Date']]
    verdict = r[idx['Step1Verdict']]
    if d is None or verdict not in ('BUY','ASYM'):
        continue
    if isinstance(d, datetime):
        d2 = d.date()
    elif isinstance(d, date):
        d2 = d
    else:
        try:
            d2 = datetime.strptime(str(d), '%m/%d/%Y').date()
        except:
            try:
                d2 = datetime.strptime(str(d), '%Y-%m-%d').date()
            except:
                continue
    if d2 >= cutoff:
        rows.append((str(d2), r[idx['Symbol']], r[idx['SignalSource']], verdict, r[idx.get('Traded')], r[idx.get('RecheckStatus')] if 'RecheckStatus' in idx else None))

with open(r'C:\Users\Trader\AI-Agent-Learning-Hub\Agentic-Hub-Governance\verify\buy_since_cutover_20260817_094033.txt', 'w', encoding='utf-8') as f:
    f.write(f'Total BUY/ASYM rows since 2026-07-24: {len(rows)}\n')
    for x in rows:
        f.write(' | '.join(str(v) for v in x) + '\n')
print('done', len(rows))
