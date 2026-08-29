import openpyxl
wb = openpyxl.load_workbook(r"C:\Users\Trader\AI-Agent-Learning-Hub\projects\P_115_BuytheDipTradingSystem\data\P_115_TrackerDashboard_V3.xlsx", data_only=True)
ws = wb["Tracker Log"]
header = [c.value for c in ws[1]]
idx = {}
for i, h in enumerate(header):
    if h:
        idx[h] = i
rows = list(ws.iter_rows(min_row=2, values_only=True))
real = [r for r in rows if r[idx["Symbol"]] is not None]
lines = []
lines.append("Total non-blank rows: " + str(len(real)))
lines.append("Last 15 non-blank rows:")
for r in real[-15:]:
    lines.append("Date=" + str(r[idx["Date"]]) + " Symbol=" + str(r[idx["Symbol"]]) + " Source=" + str(r[idx["SignalSource"]]) + " Verdict=" + str(r[idx["Step1Verdict"]]))
outp = r"C:\Users\Trader\AI-Agent-Learning-Hub\Agentic-Hub-Governance\verify\diag_output2.txt"
with open(outp, "w", encoding="utf-8") as f:
    f.write(chr(10).join(lines))
print("done")
