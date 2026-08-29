"""
PEH run_this — P_025 Formula Gaps (Cost Basis + Invested Value)
Timestamp: 20260821_225833

1. Confirms new modules / Daily_Invested support are present.
2. Re-runs full build (populates Daily_Invested).
3. Re-runs format_analytics (versioned workbook).
4. Verifies Cost Basis and Invested Value formulas are no longer hard-coded 0.

Prints PASS only when all checks succeed.
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

HUB_ROOT = Path(r"C:\Users\Trader\AI-Agent-Learning-Hub")
PROJECT_ROOT = HUB_ROOT / "projects" / "P_025_AJZ_Institutional_Portfolio_Tracker"
PYTHON_ROOT = PROJECT_ROOT / "python"
SOURCE_WB = PROJECT_ROOT / "output" / "P_025_Portfolio_BUILT.xlsx"


def write_done(status: str, exit_code: int) -> None:
    done_path = Path(__file__).with_suffix(Path(__file__).suffix + ".done")
    done_path.write_text(
        f"status={status}\nexit_code={exit_code}\ntimestamp={datetime.now().isoformat()}\n",
        encoding="utf-8",
    )


def fail(msg: str) -> None:
    print(f"FAIL: {msg}")
    write_done("FAIL", 1)
    sys.exit(1)


def main() -> int:
    print("=" * 60)
    print("P_025 Formula Gaps — 20260821_225833")
    print("=" * 60)

    if not PYTHON_ROOT.exists():
        fail(f"python\\ missing: {PYTHON_ROOT}")

    # Presence checks
    for rel in (
        "domain/trade_processor.py",
        "schemas.py",
        "infrastructure/excel_writer.py",
        "infrastructure/analytics_sheets.py",
        "application/build_portfolio.py",
    ):
        if not (PYTHON_ROOT / rel).exists():
            fail(f"Missing {rel}. Extract updated zip over python\\ and re-run.")
    print("  OK  Required modules present")

    if str(PYTHON_ROOT) not in sys.path:
        sys.path.insert(0, str(PYTHON_ROOT))

    # 1. Full build (creates Daily_Invested)
    print("\n=== Full build ===")
    try:
        from application.build_portfolio import run_full_build
        snap = run_full_build()
        print(f"  OK  Trades={len(snap.trades)}  Daily_Invested rows={len(snap.daily_invested)}")
        if snap.trades and len(snap.daily_invested) == 0:
            fail("Daily_Invested is empty despite having trades")
    except Exception as exc:
        fail(f"build raised: {type(exc).__name__}: {exc}")

    if not SOURCE_WB.exists():
        fail(f"Source workbook missing after build: {SOURCE_WB}")

    # Confirm Daily_Invested sheet exists
    from openpyxl import load_workbook
    wb = load_workbook(SOURCE_WB, read_only=True, data_only=False)
    if "Daily_Invested" not in wb.sheetnames:
        wb.close()
        fail("Daily_Invested sheet missing from Data Lake workbook")
    di_rows = wb["Daily_Invested"].max_row
    print(f"  OK  Daily_Invested max_row={di_rows}")
    wb.close()

    # 2. Format analytics
    print("\n=== format_analytics ===")
    try:
        from application.format_analytics import run_format_analytics
        dest = run_format_analytics(source_path=SOURCE_WB)
        print(f"  OK  Versioned workbook → {dest}")
    except Exception as exc:
        fail(f"format_analytics raised: {type(exc).__name__}: {exc}")

    # 3. Verify formulas are not hard-coded 0
    print("\n=== Formula checks ===")
    wb = load_workbook(dest, read_only=True, data_only=False)

    # Positions Cost Basis (col E, first data row = 3)
    pos = wb["Positions"]
    cost_cell = pos.cell(3, 5).value
    print(f"  Positions E3 (Cost Basis) = {cost_cell!r}")
    if cost_cell in (0, "0", None):
        fail("Positions Cost Basis still hard-coded 0")
    if not (isinstance(cost_cell, str) and "Trade_Log" in cost_cell):
        fail(f"Positions Cost Basis does not reference Trade_Log: {cost_cell!r}")
    print("  OK  Cost Basis formula references Trade_Log")

    # Equity_Curve Invested Value (col C, first data row = 2)
    eq = wb["Equity_Curve"]
    inv_cell = eq.cell(2, 3).value
    print(f"  Equity_Curve C2 (Invested) = {inv_cell!r}")
    if inv_cell in (0, "0", None):
        fail("Equity_Curve Invested Value still hard-coded 0")
    if not (isinstance(inv_cell, str) and "Daily_Invested" in inv_cell):
        fail(f"Invested Value does not reference Daily_Invested: {inv_cell!r}")
    print("  OK  Invested Value formula references Daily_Invested")

    wb.close()

    print("\n" + "=" * 60)
    print("PASS")
    print("=" * 60)
    write_done("PASS", 0)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
