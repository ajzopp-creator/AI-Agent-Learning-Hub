"""
PEH run_this — P_025 Analytics Expansion (Risk / Sector / Correlation)
Timestamp: 20260822_151029

1. Confirms new modules present and under 300 lines.
2. Runs format_analytics on existing Data Lake workbook.
3. Verifies Risk_Metrics, Sector_Exposure, Correlation are populated.

Prints PASS only when all checks succeed.
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

HUB_ROOT = Path(r"C:\Users\Trader\AI-Agent-Learning-Hub")
PROJECT_ROOT = HUB_ROOT / "projects" / "P_025_AJZ_Institutional_Portfolio_Tracker"
PYTHON_ROOT = PROJECT_ROOT / "python"
SOURCE_WB = PROJECT_ROOT / "output" / "P_025_Portfolio_BUILT.xlsx"
MAX_LINES = 300


def write_done(status: str, exit_code: int) -> None:
    done_path = Path(__file__).with_suffix(Path(__file__).suffix + ".done")
    done_path.write_text(
        f"status={status}\nexit_code={exit_code}\ntimestamp={datetime.now().isoformat()}\n",
        encoding="utf-8",
    )


def fail(msg: str) -> None:
    print(f"FAIL: {msg}")
    write_done("FAIL", 1)
    sys.exit(1)


def main() -> int:
    print("=" * 60)
    print("P_025 Analytics Expansion — 20260822_151029")
    print("=" * 60)

    if not PYTHON_ROOT.exists():
        fail(f"python\\ missing: {PYTHON_ROOT}")
    if not SOURCE_WB.exists():
        fail(f"Source workbook missing: {SOURCE_WB}")

    required = [
        "infrastructure/analytics_sheets.py",
        "infrastructure/analytics_exposure.py",
        "application/format_analytics.py",
        "domain/formula_templates.py",
    ]
    print("\n=== Line counts ===")
    for rel in required:
        path = PYTHON_ROOT / rel
        if not path.exists():
            fail(f"Missing {rel}. Extract updated zip over python\\ and re-run.")
        lines = len(path.read_text(encoding="utf-8").splitlines())
        print(f"  {'OK' if lines <= MAX_LINES else 'FAIL'}  {rel}: {lines}")
        if lines > MAX_LINES:
            fail(f"{rel} exceeds {MAX_LINES}-line limit")

    if str(PYTHON_ROOT) not in sys.path:
        sys.path.insert(0, str(PYTHON_ROOT))

    print("\n=== format_analytics ===")
    try:
        from application.format_analytics import run_format_analytics
        dest = run_format_analytics(source_path=SOURCE_WB)
        print(f"  OK  → {dest}")
    except Exception as exc:
        fail(f"format_analytics raised: {type(exc).__name__}: {exc}")

    print("\n=== Sheet checks ===")
    from openpyxl import load_workbook
    wb = load_workbook(dest, read_only=True, data_only=False)

    # Risk_Metrics — Ann Return should not be TBD
    rm = wb["Risk_Metrics"]
    ann_ret = rm.cell(3, 2).value
    print(f"  Risk_Metrics B3 (Ann Return) = {ann_ret!r}")
    if ann_ret in (None, "TBD", 0, "0"):
        fail("Risk_Metrics Ann Return still TBD/empty")
    if not (isinstance(ann_ret, str) and "Equity_Curve" in ann_ret):
        fail("Ann Return does not reference Equity_Curve")

    se = wb["Sector_Exposure"]
    if se.max_row < 3:
        fail(f"Sector_Exposure appears empty (max_row={se.max_row})")
    print(f"  OK  Sector_Exposure max_row={se.max_row}")

    corr = wb["Correlation"]
    if corr.max_row < 4:
        fail(f"Correlation appears empty (max_row={corr.max_row})")
    print(f"  OK  Correlation max_row={corr.max_row}")

    wb.close()

    print("\n" + "=" * 60)
    print("PASS")
    print("=" * 60)
    write_done("PASS", 0)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
