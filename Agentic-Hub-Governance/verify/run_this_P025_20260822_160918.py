"""
PEH run_this — P_025 True Average Cost
Timestamp: 20260822_160918

1. Full build (populates Cost_Basis sheet).
2. format_analytics (Positions VLOOKUP).
3. Verifies Cost_Basis sheet + Positions E3 references Cost_Basis.

Prints PASS only when all checks succeed.
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

HUB_ROOT = Path(r"C:\Users\Trader\AI-Agent-Learning-Hub")
PROJECT_ROOT = HUB_ROOT / "projects" / "P_025_AJZ_Institutional_Portfolio_Tracker"
PYTHON_ROOT = PROJECT_ROOT / "python"
SOURCE_WB = PROJECT_ROOT / "output" / "P_025_Portfolio_BUILT.xlsx"


def write_done(status: str, exit_code: int) -> None:
    done_path = Path(__file__).with_suffix(Path(__file__).suffix + ".done")
    done_path.write_text(
        f"status={status}\nexit_code={exit_code}\ntimestamp={datetime.now().isoformat()}\n",
        encoding="utf-8",
    )


def fail(msg: str) -> None:
    print(f"FAIL: {msg}")
    write_done("FAIL", 1)
    sys.exit(1)


def main() -> int:
    print("=" * 60)
    print("P_025 True Average Cost — 20260822_160918")
    print("=" * 60)

    if not PYTHON_ROOT.exists():
        fail(f"python\\ missing: {PYTHON_ROOT}")

    for rel in (
        "domain/trade_processor.py",
        "schemas.py",
        "infrastructure/excel_writer.py",
        "application/build_portfolio.py",
        "infrastructure/analytics_sheets.py",
    ):
        if not (PYTHON_ROOT / rel).exists():
            fail(f"Missing {rel}. Extract updated zip over python\\ and re-run.")
    print("  OK  Modules present")

    if str(PYTHON_ROOT) not in sys.path:
        sys.path.insert(0, str(PYTHON_ROOT))

    print("\n=== Full build ===")
    try:
        from application.build_portfolio import run_full_build
        snap = run_full_build()
        print(f"  OK  Trades={len(snap.trades)}  Cost_Basis rows={len(snap.cost_basis)}")
        if snap.trades and len(snap.cost_basis) == 0:
            fail("Cost_Basis empty despite trades (unexpected for long book)")
    except Exception as exc:
        fail(f"build raised: {type(exc).__name__}: {exc}")

    from openpyxl import load_workbook
    wb = load_workbook(SOURCE_WB, read_only=True, data_only=False)
    if "Cost_Basis" not in wb.sheetnames:
        wb.close()
        fail("Cost_Basis sheet missing from Data Lake")
    print(f"  OK  Cost_Basis max_row={wb['Cost_Basis'].max_row}")
    wb.close()

    print("\n=== format_analytics ===")
    try:
        from application.format_analytics import run_format_analytics
        dest = run_format_analytics(source_path=SOURCE_WB)
        print(f"  OK  → {dest}")
    except Exception as exc:
        fail(f"format_analytics raised: {type(exc).__name__}: {exc}")

    wb = load_workbook(dest, read_only=True, data_only=False)
    cell = wb["Positions"].cell(3, 5).value
    print(f"  Positions E3 = {cell!r}")
    if not (isinstance(cell, str) and "Cost_Basis" in cell):
        fail("Positions Cost Basis does not VLOOKUP Cost_Basis")
    print("  OK  Cost Basis references Cost_Basis sheet")
    wb.close()

    print("\n" + "=" * 60)
    print("PASS")
    print("=" * 60)
    write_done("PASS", 0)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
