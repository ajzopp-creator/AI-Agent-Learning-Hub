"""
FILE: P_300_vantagepoint_batch_convert_v4.py
VERSION: 4.0
DATE: 2026-05-06
DESCRIPTION: Converts VantagePoint XLSX exports to CSV. Enforces YYYY-MM-DD date 
             normalization at the source. Routes 'Pattern' files to the historical 
             hub and all other files to the live evaluation hub. Archives the 
             original XLSX files.
"""
import os
import zipfile
import datetime
import csv
from openpyxl import load_workbook

# ---------------------------------------------------------
# CONFIGURATION & PATHS
# ---------------------------------------------------------
# Source Folder (Where VP exports the Excel files)
WORKING_FOLDER = r"D:\OneDrive\Documents\AJZStrategiesLLC\P_300_Vantage Point Up Trend Pattern Recognition"

# Destination Folders (AI Agent Hub)
PATTERN_DIR = r"C:\Users\Trader\AI-Agent-Learning-Hub\projects\P_300_Vantage_Point_Pattern_Recognition\data\historical_patterns"
EVAL_DIR = r"C:\Users\Trader\AI-Agent-Learning-Hub\projects\P_300_Vantage_Point_Pattern_Recognition\data\live"

# Archive Path
EXCEL_ARCHIVE_ZIP = os.path.join(WORKING_FOLDER, "P_300_ArchiveVP_excel.zip")

# Ensure destination folders exist
os.makedirs(PATTERN_DIR, exist_ok=True)
os.makedirs(EVAL_DIR, exist_ok=True)

# ---------------------------------------------------------
# NORMALIZATION LOGIC
# ---------------------------------------------------------
def clean_cell(val):
    """Detects datetime objects from Excel and forces YYYY-MM-DD format."""
    if isinstance(val, datetime.datetime):
        return val.strftime('%Y-%m-%d')
    return val

def convert_excel_to_csv(excel_path, csv_path):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    with open(csv_path, "w", newline="", encoding="utf-8") as f_out:
        writer = csv.writer(f_out)
        for row in ws.iter_rows(values_only=True):
            # Apply strict formatting to all cells in the row
            cleaned_row = [clean_cell(cell) for cell in row]
            
            # Only write the row if it's not entirely empty
            if any(c is not None and str(c).strip() != "" for c in cleaned_row):
                writer.writerow(cleaned_row)

# ---------------------------------------------------------
# MAIN EXECUTION
# ---------------------------------------------------------
def process_all_files():
    converted_files = []
    
    # 1. Gather Excel files
    files = [f for f in os.listdir(WORKING_FOLDER) if f.lower().endswith(".xlsx") and not f.startswith("~$")]

    if not files:
        print("No Excel files found in the working folder.")
        return

    # 2. Open Archive Zip (Append mode if exists, otherwise write)
    zip_mode = "a" if os.path.exists(EXCEL_ARCHIVE_ZIP) else "w"
    with zipfile.ZipFile(EXCEL_ARCHIVE_ZIP, zip_mode, zipfile.ZIP_DEFLATED) as archive:
        
        for filename in files:
            input_path = os.path.join(WORKING_FOLDER, filename)
            base, _ = os.path.splitext(filename)
            
            # 3. Routing Logic
            if "pattern" in base.lower():
                output_csv = os.path.join(PATTERN_DIR, base + ".csv")
                file_type = "PATTERN"
            else:
                output_csv = os.path.join(EVAL_DIR, base + ".csv")
                file_type = "EVALUATION"

            print(f"[{file_type}] Converting: {filename} -> Routed to Agent Hub")
            
            # 4. Convert and Normalize
            convert_excel_to_csv(input_path, output_csv)
            converted_files.append((file_type, base + ".csv"))

            # 5. Archive and Cleanup
            archive.write(input_path, arcname=filename)
            os.remove(input_path)

    # 6. Summary Output
    print("\n-----------------------------------------")
    print(" P_300 CONVERSION & ROUTING SUMMARY")
    print("-----------------------------------------")
    patterns_count = sum(1 for t, _ in converted_files if t == "PATTERN")
    evals_count = sum(1 for t, _ in converted_files if t == "EVALUATION")
    
    print(f"Total Processed: {len(converted_files)}")
    print(f"  -> {patterns_count} routed to historical_patterns")
    print(f"  -> {evals_count} routed to live")
    print(f"Original XLSX files secured in: {EXCEL_ARCHIVE_ZIP}")
    print("-----------------------------------------\n")

if __name__ == "__main__":
    process_all_files()