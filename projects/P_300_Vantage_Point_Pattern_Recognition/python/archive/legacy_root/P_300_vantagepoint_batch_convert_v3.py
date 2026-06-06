import os
import zipfile
import datetime
import csv
from openpyxl import load_workbook

# ---------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------

# Correct working folder on D:
WORKING_FOLDER = r"D:\OneDrive\Documents\AJZStrategiesLLC\P_300_Vantage Point Up Trend Pattern Recognition"

# Both input and output happen in this folder
INPUT_FOLDER = WORKING_FOLDER
OUTPUT_FOLDER = WORKING_FOLDER

# ---------------------------------------------------------
# MONTHLY ZIP NAME
# ---------------------------------------------------------

today = datetime.date.today()
month_name = today.strftime("%B")
year = today.strftime("%Y")

zip_name = f"P_300_{month_name}_{year}_MTD.zip"
zip_path = os.path.join(WORKING_FOLDER, zip_name)

zip_mode = "a" if os.path.exists(zip_path) else "w"
monthly_zip = zipfile.ZipFile(zip_path, zip_mode, zipfile.ZIP_DEFLATED)

# ---------------------------------------------------------
# EXCEL → CSV CONVERSION
# ---------------------------------------------------------

def convert_excel_to_csv(excel_path, csv_path):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    with open(csv_path, "w", newline="", encoding="utf-8") as f_out:
        writer = csv.writer(f_out)

        for row in ws.iter_rows(values_only=True):
            writer.writerow(list(row) if row else [])

# ---------------------------------------------------------
# MAIN PROCESSING LOOP
# ---------------------------------------------------------

def process_all_files():
    converted_files = []

    # Only process Excel files
    files = [f for f in os.listdir(INPUT_FOLDER)
             if f.lower().endswith(".xlsx")]

    if not files:
        print("No Excel files found in the working folder.")
        input("\nPress Enter to exit...")
        return

    for filename in files:
        input_path = os.path.join(INPUT_FOLDER, filename)

        # Skip ZIP files or anything else
        if filename.lower().endswith(".zip"):
            continue

        base, _ = os.path.splitext(filename)
        output_csv = os.path.join(OUTPUT_FOLDER, base + ".csv")

        print(f"Converting Excel: {filename} → {base}.csv")

        convert_excel_to_csv(input_path, output_csv)
        converted_files.append(base + ".csv")

        print(f"Archiving original Excel file into {zip_name}")
        monthly_zip.write(input_path, arcname=filename)

        os.remove(input_path)

    print("\nAll Excel files processed successfully.")
    print(f"Monthly archive updated: {zip_path}")

    print("\n-----------------------------------------")
    print(" Conversion Summary")
    print("-----------------------------------------")
    for f in converted_files:
        print(f"  ✔ {f}")
    print("-----------------------------------------")

    input("\nPress Enter to exit...")


# ---------------------------------------------------------
# EXECUTION
# ---------------------------------------------------------

if __name__ == "__main__":
    process_all_files()
    monthly_zip.close()
