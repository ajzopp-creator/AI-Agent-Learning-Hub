"""
FILE: vp_xlsx_reader.py
VERSION: 1.4
DATE: 2026-06-17
AUTHOR: Anthony Zoppi + Claude
LAYER: infrastructure
DESCRIPTION:
    Parses VantagePoint History Grid XLSX exports. Two public entry points,
    one per pipeline:

        - parse_pattern_file()  Pipeline A -- Pattern_YYYYMMDD_YYYYMMDD_SYMBOL.xlsx
        - parse_live_file()     Pipeline B -- History Grid (SYMBOL).xlsx

    Both share manifest load, sheet selection, column-count check, header
    verification, and bar extraction; only filename parsing differs. Live
    files are the original VP exports under their default names; Pattern
    files are the same content renamed for archival capture.

    Pure I/O -- no normalization, no window slicing, no DB writes. The
    application layer selects the LAUNCH-anchor setup window + forward-label
    bars (Pipeline A) or the most-recent N-bar candidate window (Pipeline B).

    Filename conventions:
        Pattern_YYYYMMDD_YYYYMMDD_SYMBOL.xlsx   (Pipeline A)
        History Grid (SYMBOL).xlsx              (Pipeline B)

    Sheet convention:
        Workbook contains one sheet named after the symbol; sheet is
        targeted by name (not wb.active) so save-state or accidental tab
        renames cannot silently retarget the parse. Mismatch fails fast.

    Manifest:
        parameters/ingest_manifest.json, validated at load time via
        IngestManifest (schemas.py). Column index -> pattern_bars field
        mapping is the single source of truth for VP -> catalog
        translation. Updates to VP column layout are isolated to the
        manifest -- no code change required.

    Validation order (fail-fast, in order):
        1. Filename matches regex
        2. XLSX and manifest files exist
        3. Sheet named after symbol exists in workbook
        4. Column count matches manifest (when strict_column_count=true)
        5. Header text matches manifest (when verify_header_text=true)
        6. Every mapped cell coerces cleanly to its declared type
        7. VPBarRaw model validates each bar (OHLC > 0, high >= low, etc.)
        8. PatternFileParse validates the full set (>= 60 bars, ascending)

CHANGELOG:
    - 2026-06-17 v1.4: Fixed v1.3 regression -- filename.upper() was uppercasing
      the ENTIRE filename before the regex match, breaking the literal
      'Pattern_' / '.xlsx' portions and causing 100% ingest failure since
      v1.3 shipped (caught: 18/18 AddPattern run failed 2026-06-17 09:25).
      Regex now compiled with re.IGNORECASE; only the captured symbol group
      is uppercased when building PatternFileMetadata. O-009 intent (RCl ->
      RCL) preserved without mangling the rest of the filename.
    - 2026-06-16 v1.3: Auto-uppercase filename before regex match in
      _parse_filename(). Prevents case-mismatch ValueError on files saved
      with lowercase symbol (e.g. RCl -> RCL). O-009.
    - 2026-05-20 v1.2: Updated _verify_header_text to accept header_sub_alt
      when present in the manifest ColumnMapEntry. Allows both old and new
      VP triple_cross sub-header formats ('Short' vs 'Triple Cross Short')
      to pass validation without toggling the manifest. Primary header_sub
      reflects the current VP version; header_sub_alt covers legacy exports.
    - 2026-05-17 v1.1: Stage 6 file #8 prep. Added Pipeline B entry point
      parse_live_file() for History Grid (SYMBOL).xlsx exports.
    - 2026-05-15 v1.0: Stage 4 file #5. Targets sheet by symbol name;
      manifest-driven column mapping; header-text verification per M-014.
"""
from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

_PYTHON_DIR = Path(__file__).resolve().parent.parent
if str(_PYTHON_DIR) not in sys.path:
    sys.path.insert(0, str(_PYTHON_DIR))

from config import LOG_FORMAT, LOG_LEVEL, PARAMETERS_DIR  # noqa: E402
from schemas import (  # noqa: E402
    IngestManifest,
    PatternFileMetadata,
    PatternFileParse,
    VPBarRaw,
)

logger = logging.getLogger(__name__)

_FILENAME_PATTERN = re.compile(
    r"^Pattern_(?P<start>\d{8})_(?P<end>\d{8})_(?P<symbol>[A-Za-z][A-Za-z0-9_]{0,11})\.xlsx$",
    re.IGNORECASE,
)
_LIVE_FILENAME_PATTERN = re.compile(
    r"^History Grid \((?P<symbol>[A-Z][A-Z0-9_]{0,11})\)\.xlsx$"
)
_FILE_DATE_FORMAT = "%Y%m%d"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _load_manifest(manifest_path: Path) -> IngestManifest:
    logger.info("Loading manifest: %s", manifest_path)
    with manifest_path.open("r", encoding="utf-8") as f:
        raw = json.load(f)
    return IngestManifest.model_validate(raw)


def _parse_filename(filename: str) -> PatternFileMetadata:
    m = _FILENAME_PATTERN.match(filename)
    if not m:
        raise ValueError(
            f"Filename does not match Pattern_YYYYMMDD_YYYYMMDD_SYMBOL.xlsx: "
            f"got {filename!r}"
        )
    try:
        start_d = datetime.strptime(m.group("start"), _FILE_DATE_FORMAT).date()
        end_d = datetime.strptime(m.group("end"), _FILE_DATE_FORMAT).date()
    except ValueError as e:
        raise ValueError(f"Invalid date in filename {filename!r}: {e}") from e
    return PatternFileMetadata(
        filename=filename,
        symbol=m.group("symbol").upper(),  # v1.4: normalize symbol case only (O-009)
        pattern_start_date=start_d,
        pattern_end_date=end_d,
    )


def _parse_live_filename(filename: str) -> str:
    m = _LIVE_FILENAME_PATTERN.match(filename)
    if not m:
        raise ValueError(
            f"Filename does not match 'History Grid (SYMBOL).xlsx': "
            f"got {filename!r}"
        )
    return m.group("symbol")


def _select_sheet(workbook, symbol: str):
    if symbol not in workbook.sheetnames:
        raise ValueError(
            f"Sheet {symbol!r} not found in workbook. "
            f"Available sheets: {workbook.sheetnames}"
        )
    return workbook[symbol]


def _validate_column_count(sheet, manifest: IngestManifest) -> None:
    if not manifest.validation_rules.strict_column_count:
        return
    expected = manifest.source_format.expected_column_count
    actual = sheet.max_column
    if actual != expected:
        raise ValueError(
            f"Column count mismatch in sheet {sheet.title!r}: "
            f"expected {expected}, got {actual}. Manifest may be stale "
            f"or VP export format has drifted."
        )


def _verify_header_text(sheet, manifest: IngestManifest) -> None:
    """Verify each mapped column's header_top and header_sub against the
    manifest. When a ColumnMapEntry has header_sub_alt set, either value
    is accepted -- this covers VP version transitions where the sub-header
    text changes but the data is identical (e.g. triple_cross columns in
    VP v10.0.2504.0114)."""
    if not manifest.validation_rules.verify_header_text:
        return

    header_top_row = 1
    header_sub_row = 2 if manifest.source_format.header_rows >= 2 else None

    mismatches: list[str] = []
    for col_idx, entry in manifest.column_mapping.items():
        actual_top = sheet.cell(row=header_top_row, column=col_idx + 1).value
        if actual_top != entry.header_top:
            mismatches.append(
                f"col {col_idx} ({entry.field}): header_top expected "
                f"{entry.header_top!r}, got {actual_top!r}"
            )
        if header_sub_row is not None:
            actual_sub = sheet.cell(row=header_sub_row, column=col_idx + 1).value
            accepted = {entry.header_sub}
            if entry.header_sub_alt is not None:
                accepted.add(entry.header_sub_alt)
            if actual_sub not in accepted:
                mismatches.append(
                    f"col {col_idx} ({entry.field}): header_sub expected "
                    f"{entry.header_sub!r}, got {actual_sub!r}"
                )

    if mismatches and manifest.validation_rules.raise_on_header_mismatch:
        joined = "\n  ".join(mismatches)
        raise ValueError(
            f"Header text mismatches against manifest:\n  {joined}\n"
            f"Either VP changed the export or the manifest is stale. "
            f"Update parameters/ingest_manifest.json."
        )


def _coerce_cell(value, col_type: str, col_idx: int, row_num: int):
    if value is None:
        raise ValueError(
            f"Empty cell at row {row_num} col {col_idx} (expected {col_type})"
        )
    if col_type == "date":
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
            raise ValueError(
                f"Cannot parse date string {value!r} at row {row_num} col {col_idx}"
            )
        raise TypeError(
            f"Unexpected date cell type {type(value).__name__} "
            f"at row {row_num} col {col_idx}"
        )
    if col_type == "float":
        try:
            return float(value)
        except (TypeError, ValueError) as e:
            raise ValueError(
                f"Cannot coerce {value!r} to float at row {row_num} col {col_idx}: {e}"
            ) from e
    raise ValueError(f"Unknown col_type {col_type!r} in manifest")


def _extract_bars(sheet, manifest: IngestManifest) -> list[VPBarRaw]:
    start_row = manifest.source_format.data_start_row_index + 1
    end_row = sheet.max_row
    bars: list[VPBarRaw] = []
    for row_num in range(start_row, end_row + 1):
        bar_kwargs: dict = {}
        for col_idx, entry in manifest.column_mapping.items():
            raw = sheet.cell(row=row_num, column=col_idx + 1).value
            bar_kwargs[entry.field] = _coerce_cell(raw, entry.type, col_idx, row_num)
        bars.append(VPBarRaw(**bar_kwargs))
    if manifest.source_format.date_order == "descending":
        bars.reverse()
    return bars


# ---------------------------------------------------------------------------
# Public entry points
# ---------------------------------------------------------------------------

def parse_pattern_file(
    xlsx_path: Path,
    manifest_path: Path | None = None,
) -> PatternFileParse:
    if manifest_path is None:
        manifest_path = PARAMETERS_DIR / "ingest_manifest.json"
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")
    if not manifest_path.exists():
        raise FileNotFoundError(f"Manifest not found: {manifest_path}")

    logger.info("Parsing %s", xlsx_path.name)
    metadata = _parse_filename(xlsx_path.name)
    manifest = _load_manifest(manifest_path)
    workbook = load_workbook(xlsx_path, data_only=True)
    try:
        sheet = _select_sheet(workbook, metadata.symbol)
        _validate_column_count(sheet, manifest)
        _verify_header_text(sheet, manifest)
        bars = _extract_bars(sheet, manifest)
    finally:
        workbook.close()

    logger.info("Parsed %d bars from %s", len(bars), xlsx_path.name)
    return PatternFileParse(metadata=metadata, bars=bars)


def parse_live_file(
    xlsx_path: Path,
    manifest_path: Path | None = None,
) -> tuple[str, list[VPBarRaw]]:
    if manifest_path is None:
        manifest_path = PARAMETERS_DIR / "ingest_manifest.json"
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")
    if not manifest_path.exists():
        raise FileNotFoundError(f"Manifest not found: {manifest_path}")

    logger.info("Parsing live file %s", xlsx_path.name)
    symbol = _parse_live_filename(xlsx_path.name)
    manifest = _load_manifest(manifest_path)
    workbook = load_workbook(xlsx_path, data_only=True)
    try:
        sheet = _select_sheet(workbook, symbol)
        _validate_column_count(sheet, manifest)
        _verify_header_text(sheet, manifest)
        bars = _extract_bars(sheet, manifest)
    finally:
        workbook.close()

    logger.info("Parsed %d bars from %s", len(bars), xlsx_path.name)
    return symbol, bars


# ---------------------------------------------------------------------------
# CLI self-test
# ---------------------------------------------------------------------------

def _selftest(xlsx_path_str: str) -> int:
    logging.basicConfig(level=LOG_LEVEL, format=LOG_FORMAT, stream=sys.stdout)
    xlsx_path = Path(xlsx_path_str)
    try:
        parse = parse_pattern_file(xlsx_path)
    except Exception as e:
        print(f"\nPARSE FAILED: {type(e).__name__}: {e}")
        return 1

    md = parse.metadata
    print(f"\nFile:    {md.filename}")
    print(f"Symbol:  {md.symbol}")
    print(f"Pattern: {md.pattern_start_date} -> {md.pattern_end_date}")
    print(
        f"Bars:    {len(parse.bars)} "
        f"(file range {parse.bars[0].bar_date} -> {parse.bars[-1].bar_date})"
    )
    print("\nFirst 3 bars:")
    for b in parse.bars[:3]:
        print(
            f"  {b.bar_date}  O={b.open:7.2f}  H={b.high:7.2f}  "
            f"L={b.low:7.2f}  C={b.close:7.2f}  V={b.volume:>12,.0f}"
        )
    print("\nLast 3 bars:")
    for b in parse.bars[-3:]:
        print(
            f"  {b.bar_date}  O={b.open:7.2f}  H={b.high:7.2f}  "
            f"L={b.low:7.2f}  C={b.close:7.2f}  V={b.volume:>12,.0f}"
        )
    return 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Parse one VP Pattern XLSX as a self-test."
    )
    parser.add_argument("--xlsx", required=True)
    args = parser.parse_args()
    sys.exit(_selftest(args.xlsx))
