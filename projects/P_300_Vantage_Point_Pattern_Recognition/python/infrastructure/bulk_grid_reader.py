"""
FILE: infrastructure/bulk_grid_reader.py
VERSION: 1.2
DATE: 2026-07-09
AUTHOR: Anthony Zoppi + Claude
LAYER: infrastructure
DESCRIPTION:
    Parses bulk History Grid XLSX exports for WO-P300-E2.001 Pipeline
    A-Bulk. Separate module from infrastructure/vp_xlsx_reader.py per
    the WO's own instruction ("New reader module required -- do NOT
    reuse vp_xlsx_reader") -- the bulk raw shape genuinely differs
    (text neural_index, price-level Triple Cross), not just the
    filename convention.

    Pure I/O -- no detection, no window slicing, no DB writes. The
    domain layer (bulk_windowing.py, bulk_pattern_detector.py) consumes
    the parsed bars. CLI self-test lives in the companion
    bulk_grid_reader_selftest.py, split out at v1.1 to keep this file
    under the 300-line hard limit (see changelog).

    Filename convention (operator-typed at export time; VP defaults to
    "History Grid (SYMBOL).xlsx" and the operator renames):
        <years>[I]_Pattern_<SYMBOL>.xlsx
        e.g. 10_Pattern_SPY.xlsx, 5_Pattern_AAPL.xlsx, 1I_Pattern_BP.xlsx
    years = arbitrary export window length, no minimum or fixed set.
    I = optional flag recording an IntelliScan-routed export; verified
    byte-identical content to VP-direct at matched windows (2026-07-08)
    -- provenance only, never a parsing branch.

    Sheet convention: workbook contains one sheet named after the
    symbol, targeted by name (same fail-fast rationale as
    vp_xlsx_reader.py).

    Manifest: parameters/bulk_ingest_manifest.json, validated via the
    SHARED IngestManifest model (schemas.py v2.3 widened
    ColumnMapEntry.type to include "text" specifically so this manifest
    could declare the Neural Index column without a duplicate manifest
    schema in schemas_bulk.py -- PEH-verified safe against live
    Pipeline A, 2026-07-08, 5/5 checks PASS).

    Validation order (fail-fast, in order):
        1. Filename matches regex
        2. XLSX and manifest files exist
        3. Sheet named after symbol exists in workbook
        4. Column count matches manifest
        5. Header text matches manifest (header_sub_alt tolerant)
        6. Every mapped cell coerces cleanly to its declared type
           (date / float / text)
        7. BulkBarRaw validates each bar (OHLC > 0, high >= low, etc.)
        8. BulkPatternFileParse validates the full set (ascending, min
           bar count)

CHANGELOG:
    - 2026-07-09 v1.2: PEH-verified against real 10-year LIN data
      (2026-07-09) -- found and fixed a second real parse failure in the
      same field family as v1.1's infinity fix: ROC% (and potentially
      other predictive columns) can render as an empty string '', not
      just zero or infinity, in the pre-2021-07-14 backfill window (VP
      backfills predictions exactly 5 years -- documented in this WO's
      Phase 0 finding). Confirmed at LIN row 1853, col 13 (roc_pct).
      _coerce_cell's float branch now maps a stripped-empty string to
      0.0, matching the zero convention the other predictive columns
      already use for the same pre-backfill period. Full 2,512-bar LIN
      file parses clean end-to-end after the fix.
    - 2026-07-08 v1.1: PEH-verified against real 10-year SPY data
      (2026-07-08) -- found and fixed a real parse failure: ROC%
      (col N) can genuinely render as the string '∞' (infinity), not
      corrupted input -- confirmed at SPY row 56. _coerce_cell's float
      branch now maps '∞'/'inf'/'Infinity' (and negative forms) to
      Python float('inf')/float('-inf') explicitly, since Python's
      bare float() does not parse the unicode infinity symbol. Full
      2,512-bar SPY file now parses clean end-to-end. Split the CLI
      self-test out to bulk_grid_reader_selftest.py -- this fix pushed
      the file to 310 lines, over the 300 hard limit; the self-test
      is not part of the reader's actual API and was the natural piece
      to extract.
    - 2026-07-08 v1.0: Initial release (superseded same-day, never
      shipped standalone -- the infinity bug was caught in the same
      PEH pass that produced v1.0's line-count overage).
"""
from __future__ import annotations

import json
import logging
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

_PYTHON_DIR = Path(__file__).resolve().parent.parent
if str(_PYTHON_DIR) not in sys.path:
    sys.path.insert(0, str(_PYTHON_DIR))

from config import PARAMETERS_DIR  # noqa: E402
from schemas import IngestManifest  # noqa: E402
from schemas_bulk import (  # noqa: E402
    BulkBarRaw,
    BulkPatternFileMetadata,
    BulkPatternFileParse,
)

logger = logging.getLogger(__name__)

_FILENAME_PATTERN = re.compile(
    r"^(?P<years>\d{1,2})(?P<intelliscan>I)?_Pattern_"
    r"(?P<symbol>[A-Za-z][A-Za-z0-9_]{0,11})\.xlsx$",
    re.IGNORECASE,
)
_NEURAL_INDEX_VALUES = {"up", "down"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _load_manifest(manifest_path: Path) -> IngestManifest:
    logger.info("Loading bulk manifest: %s", manifest_path)
    with manifest_path.open("r", encoding="utf-8") as f:
        raw = json.load(f)
    return IngestManifest.model_validate(raw)


def _parse_filename(filename: str) -> BulkPatternFileMetadata:
    m = _FILENAME_PATTERN.match(filename)
    if not m:
        raise ValueError(
            f"Filename does not match <years>[I]_Pattern_<SYMBOL>.xlsx: "
            f"got {filename!r}"
        )
    return BulkPatternFileMetadata(
        filename=filename,
        symbol=m.group("symbol").upper(),
        window_years=int(m.group("years")),
        is_intelliscan_export=m.group("intelliscan") is not None,
    )


def _select_sheet(workbook, symbol: str):
    if symbol not in workbook.sheetnames:
        raise ValueError(
            f"Sheet {symbol!r} not found in workbook. "
            f"Available sheets: {workbook.sheetnames}"
        )
    return workbook[symbol]


def _validate_column_count(sheet, manifest: IngestManifest) -> None:
    if not manifest.validation_rules.strict_column_count:
        return
    expected = manifest.source_format.expected_column_count
    actual = sheet.max_column
    if actual != expected:
        raise ValueError(
            f"Column count mismatch in sheet {sheet.title!r}: "
            f"expected {expected}, got {actual}."
        )


def _verify_header_text(sheet, manifest: IngestManifest) -> None:
    """Same header_sub_alt tolerance as vp_xlsx_reader.py -- required
    here from day one since the Triple Cross sub-header wording drift
    was confirmed present across real bulk exports before this manifest
    was written (not discovered live, unlike the original live-manifest
    case)."""
    if not manifest.validation_rules.verify_header_text:
        return

    header_top_row = 1
    header_sub_row = 2 if manifest.source_format.header_rows >= 2 else None

    mismatches: list[str] = []
    for col_idx, entry in manifest.column_mapping.items():
        actual_top = sheet.cell(row=header_top_row, column=col_idx + 1).value
        if actual_top != entry.header_top:
            mismatches.append(
                f"col {col_idx} ({entry.field}): header_top expected "
                f"{entry.header_top!r}, got {actual_top!r}"
            )
        if header_sub_row is not None:
            actual_sub = sheet.cell(row=header_sub_row, column=col_idx + 1).value
            accepted = {entry.header_sub}
            if entry.header_sub_alt is not None:
                accepted.add(entry.header_sub_alt)
            if actual_sub not in accepted:
                mismatches.append(
                    f"col {col_idx} ({entry.field}): header_sub expected "
                    f"{entry.header_sub!r}, got {actual_sub!r}"
                )

    if mismatches and manifest.validation_rules.raise_on_header_mismatch:
        joined = "\n  ".join(mismatches)
        raise ValueError(
            f"Header text mismatches against bulk manifest:\n  {joined}\n"
            f"Update parameters/bulk_ingest_manifest.json."
        )


def _coerce_cell(value, col_type: str, col_idx: int, row_num: int):
    if col_type == "text":
        if value is None:
            return "unknown"
        text = str(value).strip().lower()
        return text if text in _NEURAL_INDEX_VALUES else "unknown"
    if value is None:
        raise ValueError(
            f"Empty cell at row {row_num} col {col_idx} (expected {col_type})"
        )
    if col_type == "date":
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        raise TypeError(
            f"Unexpected date cell type {type(value).__name__} "
            f"at row {row_num} col {col_idx}"
        )
    if col_type == "float":
        # Real VP output: ROC% (rate of change) divides by a near-zero
        # prior value and can genuinely render as infinity -- confirmed
        # on live SPY data (row 56, col N), not corrupted input. '∞'
        # is not a Python-parseable float literal, so it needs an
        # explicit mapping rather than falling through to the generic
        # ValueError below.
        #
        # Real VP output (2nd case, confirmed live LIN data, row 1853):
        # predictive columns (incl. ROC%) render as an empty string,
        # not just zero, in the pre-backfill window (VP backfills
        # predictions exactly 5 years -- see this WO's Phase 0 finding).
        # Blank means "no data yet" here, not corruption -- mapped to
        # 0.0 to match the zero convention the other predictive columns
        # already use for the same pre-backfill period.
        if isinstance(value, str):
            stripped = value.strip()
            if stripped in ("∞", "inf", "Infinity"):
                return float("inf")
            if stripped in ("-∞", "-inf", "-Infinity"):
                return float("-inf")
            if stripped == "":
                return 0.0
        try:
            return float(value)
        except (TypeError, ValueError) as e:
            raise ValueError(
                f"Cannot coerce {value!r} to float at row {row_num} col {col_idx}: {e}"
            ) from e
    raise ValueError(f"Unknown col_type {col_type!r} in bulk manifest")


def _extract_bars(sheet, manifest: IngestManifest) -> list[BulkBarRaw]:
    start_row = manifest.source_format.data_start_row_index + 1
    end_row = sheet.max_row
    bars: list[BulkBarRaw] = []
    for row_num in range(start_row, end_row + 1):
        bar_kwargs: dict = {}
        for col_idx, entry in manifest.column_mapping.items():
            raw = sheet.cell(row=row_num, column=col_idx + 1).value
            bar_kwargs[entry.field] = _coerce_cell(raw, entry.type, col_idx, row_num)
        bars.append(BulkBarRaw(**bar_kwargs))
    if manifest.source_format.date_order == "descending":
        bars.reverse()
    return bars


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def parse_bulk_file(
    xlsx_path: Path,
    manifest_path: Path | None = None,
) -> BulkPatternFileParse:
    if manifest_path is None:
        manifest_path = PARAMETERS_DIR / "bulk_ingest_manifest.json"
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")
    if not manifest_path.exists():
        raise FileNotFoundError(f"Manifest not found: {manifest_path}")

    logger.info("Parsing bulk file %s", xlsx_path.name)
    metadata = _parse_filename(xlsx_path.name)
    manifest = _load_manifest(manifest_path)
    workbook = load_workbook(xlsx_path, data_only=True)
    try:
        sheet = _select_sheet(workbook, metadata.symbol)
        _validate_column_count(sheet, manifest)
        _verify_header_text(sheet, manifest)
        bars = _extract_bars(sheet, manifest)
    finally:
        workbook.close()

    logger.info("Parsed %d bars from %s", len(bars), xlsx_path.name)
    return BulkPatternFileParse(metadata=metadata, bars=bars)
