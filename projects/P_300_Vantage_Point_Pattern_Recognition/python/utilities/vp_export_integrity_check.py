"""
FILE: vp_export_integrity_check.py
VERSION: 1.0
DATE: 2026-05-15
AUTHOR: Anthony Zoppi + Claude
LAYER: utilities
DESCRIPTION:
    Standalone diagnostic. Validates one VantagePoint History Grid XLSX
    export against parameters/ingest_manifest.json without performing a
    full ingest. Designed to be run immediately after any VantagePoint
    version update — before any new export is allowed into Pipeline A.

    Six independent checks, each reports PASS or FAIL with specifics:
        1. Filename convention   Pattern_YYYYMMDD_YYYYMMDD_SYMBOL.xlsx
        2. Sheet name matches    workbook contains a sheet named after
                                 the symbol parsed from the filename
        3. Column count          sheet.max_column == expected_column_count
        4. Mapped column headers every mapped column's header_top and
                                 header_sub match the manifest's literal
                                 openpyxl values (M-014, O-006)
        5. Ignored column headers every ignored column's headers also
                                 verified — drift here means VP rearranged
                                 the layout and the mapping is unsafe
                                 even though we don't extract these cells
        6. First data row coerces every mapped cell in the first data row
                                 coerces cleanly to its declared type

    Exit codes:
        0 = all checks passed; export is safe to ingest
        1 = at least one check failed; update the manifest and re-run

    Imports private validators from vp_xlsx_reader so parse logic remains
    single-source-of-truth. Importing private (_-prefixed) names is
    intentional — the two modules are tightly coupled by design.

CHANGELOG:
    - 2026-05-15 v1.0: Initial release. Bundled with vp_xlsx_reader.py v1.0
      as the post-VP-update integrity-check pair.
"""
from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

from openpyxl import load_workbook

# Allow direct execution: python utilities/vp_export_integrity_check.py --xlsx ...
_PYTHON_DIR = Path(__file__).resolve().parent.parent
if str(_PYTHON_DIR) not in sys.path:
    sys.path.insert(0, str(_PYTHON_DIR))

from config import LOG_FORMAT, PARAMETERS_DIR  # noqa: E402
from infrastructure.vp_xlsx_reader import (  # noqa: E402
    _coerce_cell,
    _load_manifest,
    _parse_filename,
)


# ─────────────────────────────────────────────────────────────────────────────
# Output formatting
# ─────────────────────────────────────────────────────────────────────────────

_LABEL_WIDTH = 32


def _format_result(label: str, passed: bool, lines: list[str]) -> str:
    """Render one check as '[PASS] Label  detail' or multi-line FAIL block."""
    tag = "[PASS]" if passed else "[FAIL]"
    if not lines:
        return f"{tag} {label}"
    head = f"{tag} {label:<{_LABEL_WIDTH}} {lines[0]}"
    if len(lines) == 1:
        return head
    indent = " " * (len(tag) + 1 + _LABEL_WIDTH + 1)
    tail = "\n".join(f"{indent}{line}" for line in lines[1:])
    return f"{head}\n{tail}"


def _print_summary(results: list[tuple[str, bool, list[str]]]) -> None:
    print()
    for label, passed, lines in results:
        print(_format_result(label, passed, lines))
    total = len(results)
    passed_count = sum(1 for _, p, _ in results if p)
    overall = "PASS" if passed_count == total else "FAIL"
    print()
    print(f"OVERALL: {overall}  ({passed_count} / {total} checks passed)")
    print()


# ─────────────────────────────────────────────────────────────────────────────
# Individual checks — each returns (passed, detail_lines)
# ─────────────────────────────────────────────────────────────────────────────

def _check_filename_convention(filename: str) -> tuple[bool, list[str]]:
    try:
        _parse_filename(filename)
        return True, [filename]
    except ValueError as e:
        return False, [str(e)]


def _check_sheet_convention(workbook, metadata):
    """Returns (sheet | None, passed, lines)."""
    if metadata is None:
        # Filename didn't parse — fall back to active sheet but record it.
        sheet = workbook.active
        return sheet, True, [
            f"filename did not parse; using wb.active = {sheet.title!r}",
            f"available sheets: {workbook.sheetnames}",
        ]
    symbol = metadata.symbol
    if symbol in workbook.sheetnames:
        return workbook[symbol], True, [f"{symbol!r}"]
    return None, False, [
        f"expected sheet {symbol!r} not found",
        f"available sheets: {workbook.sheetnames}",
    ]


def _check_column_count(sheet, manifest) -> tuple[bool, list[str]]:
    expected = manifest.source_format.expected_column_count
    actual = sheet.max_column
    if actual == expected:
        return True, [f"{actual} columns (expected {expected})"]
    return False, [f"got {actual} columns, expected {expected}"]


def _check_mapped_headers(sheet, manifest) -> tuple[bool, list[str]]:
    """Verify header_top and header_sub for every mapped column."""
    header_sub_row = 2 if manifest.source_format.header_rows >= 2 else None
    mismatches: list[str] = []
    for col_idx, entry in manifest.column_mapping.items():
        actual_top = sheet.cell(row=1, column=col_idx + 1).value
        if actual_top != entry.header_top:
            mismatches.append(
                f"col {col_idx} ({entry.field}): header_top "
                f"expected {entry.header_top!r}, got {actual_top!r}"
            )
        if header_sub_row is not None:
            actual_sub = sheet.cell(row=header_sub_row, column=col_idx + 1).value
            if actual_sub != entry.header_sub:
                mismatches.append(
                    f"col {col_idx} ({entry.field}): header_sub "
                    f"expected {entry.header_sub!r}, got {actual_sub!r}"
                )
    if not mismatches:
        return True, [f"{len(manifest.column_mapping)} columns verified"]
    return False, mismatches


def _check_ignored_headers(sheet, manifest) -> tuple[bool, list[str]]:
    """Verify header_top and header_sub for every ignored column. Drift here
    means VP rearranged the layout — even though we don't extract these
    cells, the mapping above them is no longer trustworthy."""
    header_sub_row = 2 if manifest.source_format.header_rows >= 2 else None
    mismatches: list[str] = []
    for col_idx, entry in manifest.ignored_columns.items():
        actual_top = sheet.cell(row=1, column=col_idx + 1).value
        if actual_top != entry.header_top:
            mismatches.append(
                f"col {col_idx} (ignored): header_top "
                f"expected {entry.header_top!r}, got {actual_top!r}"
            )
        if header_sub_row is not None:
            actual_sub = sheet.cell(row=header_sub_row, column=col_idx + 1).value
            if actual_sub != entry.header_sub:
                mismatches.append(
                    f"col {col_idx} (ignored): header_sub "
                    f"expected {entry.header_sub!r}, got {actual_sub!r}"
                )
    if not mismatches:
        return True, [f"{len(manifest.ignored_columns)} columns verified"]
    return False, mismatches


def _check_first_row_coerces(sheet, manifest) -> tuple[bool, list[str]]:
    """Confirm every mapped cell in the first data row coerces cleanly."""
    start_row = manifest.source_format.data_start_row_index + 1
    failures: list[str] = []
    coerced = 0
    for col_idx, entry in manifest.column_mapping.items():
        raw = sheet.cell(row=start_row, column=col_idx + 1).value
        try:
            _coerce_cell(raw, entry.type, col_idx, start_row)
            coerced += 1
        except (ValueError, TypeError) as e:
            failures.append(str(e))
    if not failures:
        return True, [f"{coerced} cells coerced cleanly"]
    return False, failures


# ─────────────────────────────────────────────────────────────────────────────
# Orchestrator
# ─────────────────────────────────────────────────────────────────────────────

def run_integrity_check(xlsx_path: Path, manifest_path: Path) -> int:
    """Run all integrity checks. Returns 0 if all pass, 1 if any fail."""
    if not xlsx_path.exists():
        print(f"FATAL: XLSX not found: {xlsx_path}")
        return 1
    if not manifest_path.exists():
        print(f"FATAL: manifest not found: {manifest_path}")
        return 1

    print(f"\nVP export integrity check")
    print(f"  XLSX:     {xlsx_path}")
    print(f"  Manifest: {manifest_path}")

    results: list[tuple[str, bool, list[str]]] = []

    passed, lines = _check_filename_convention(xlsx_path.name)
    results.append(("Filename convention", passed, lines))
    try:
        metadata = _parse_filename(xlsx_path.name)
    except ValueError:
        metadata = None

    try:
        manifest = _load_manifest(manifest_path)
    except Exception as e:
        results.append(("Manifest load", False, [f"{type(e).__name__}: {e}"]))
        _print_summary(results)
        return 1

    try:
        workbook = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        results.append(("Workbook load", False, [f"{type(e).__name__}: {e}"]))
        _print_summary(results)
        return 1

    try:
        sheet, passed, lines = _check_sheet_convention(workbook, metadata)
        results.append(("Sheet name matches symbol", passed, lines))
        if sheet is None:
            _print_summary(results)
            return 1

        for label, fn in (
            ("Column count", _check_column_count),
            ("Mapped column headers", _check_mapped_headers),
            ("Ignored column headers", _check_ignored_headers),
            ("First data row coerces", _check_first_row_coerces),
        ):
            passed, lines = fn(sheet, manifest)
            results.append((label, passed, lines))
    finally:
        workbook.close()

    _print_summary(results)
    return 0 if all(p for _, p, _ in results) else 1


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description=(
            "Check a VantagePoint History Grid XLSX against "
            "ingest_manifest.json. Run after every VP version update before "
            "resuming Pipeline A ingest."
        )
    )
    parser.add_argument(
        "--xlsx",
        required=True,
        help="Path to a VP History Grid XLSX export to check.",
    )
    parser.add_argument(
        "--manifest",
        default=None,
        help="Optional manifest path override; defaults to parameters/ingest_manifest.json.",
    )
    args = parser.parse_args()

    # WARNING level keeps the structured PASS/FAIL summary clean. Errors
    # still surface; INFO chatter from vp_xlsx_reader._load_manifest is hushed.
    logging.basicConfig(level=logging.WARNING, format=LOG_FORMAT, stream=sys.stdout)

    manifest_path = (
        Path(args.manifest) if args.manifest else PARAMETERS_DIR / "ingest_manifest.json"
    )
    sys.exit(run_integrity_check(Path(args.xlsx), manifest_path))
