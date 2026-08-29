"""
P_025 Application — Format Analytics Layer

Orchestration only: load workbook → call sheet builders → save versioned copy.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from config import WORKBOOK_PATH
from infrastructure.analytics_exposure import build_correlation, build_sector_exposure
from infrastructure.analytics_sheets import (
    apply_data_lake_formatting,
    build_dashboard,
    build_equity_curve,
    build_positions,
    build_risk_metrics,
)
from infrastructure.excel_formatter import TITLE_FONT

logger = logging.getLogger(__name__)


def _versioned_path(base: Path) -> Path:
    """Return a versioned sibling path with timestamp."""
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return base.parent / f"{base.stem}_Analytics_{ts}{base.suffix}"


def run_format_analytics(
    source_path: Path | None = None,
    output_path: Path | None = None,
) -> Path:
    """
    Open the existing Data Lake workbook, apply Analytics formatting and
    formulas, then save a versioned copy. Never overwrites the source.
    """
    source = source_path or WORKBOOK_PATH
    if not source.exists():
        raise FileNotFoundError(f"Source workbook not found: {source}")

    dest = output_path or _versioned_path(source)
    dest.parent.mkdir(parents=True, exist_ok=True)

    logger.info("Loading workbook: %s", source)
    wb = load_workbook(source)

    apply_data_lake_formatting(wb)
    build_positions(wb)
    build_equity_curve(wb)
    build_dashboard(wb)
    build_risk_metrics(wb)
    build_sector_exposure(wb)
    build_correlation(wb)

    for name in ("Geographic_Exposure", "Stress_Testing", "Investment_Theses"):
        if name in wb.sheetnames:
            ws = wb[name]
            if ws.cell(1, 1).value is None:
                ws.cell(1, 1, f"{name} — placeholder (formulas TBD)").font = TITLE_FONT

    wb.save(dest)
    logger.info("Versioned workbook saved → %s", dest)
    return dest
