"""
P_025 Infrastructure — Excel Writer

Writes (or appends) Data Lake sheets into the portfolio workbook.
Creates the workbook and all required sheets if they do not exist.
Contains no business logic — only openpyxl I/O.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from schemas import (
    CostBasisRow,
    DailyCashRow,
    DailyInvestedRow,
    DailyUnitsRow,
    MarketDataRow,
    ReferenceData,
    TradeRecord,
)

logger = logging.getLogger(__name__)

# Bloomberg-style colours
HEADER_FILL = PatternFill("solid", fgColor="1C2541")
HEADER_FONT = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Segoe UI", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def _style_header(ws, headers: list[str]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _ensure_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    # Remove the default sheet; we will create named ones
    default = wb.active
    wb.remove(default)
    return wb


def _get_or_create_sheet(wb: Workbook, name: str):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)


def write_trade_log(wb: Workbook, trades: Sequence[TradeRecord]) -> None:
    ws = _get_or_create_sheet(wb, "Trade_Log")
    headers = [
        "trade_id", "account_id", "system", "underlying_symbol", "asset_type",
        "direction", "open_date", "open_datetime", "qty", "entry_price",
        "stop_price", "risk_amount", "total_commissions", "status",
        "realized_pnl", "realized_R", "schwab_transaction_id", "notes",
    ]
    _style_header(ws, headers)

    # Clear existing data rows (keep header)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    for row_idx, t in enumerate(trades, start=2):
        values = [
            t.trade_id, t.account_id, t.system, t.underlying_symbol, t.asset_type,
            t.direction, t.open_date.isoformat(), 
            t.open_datetime.isoformat() if t.open_datetime else None,
            t.qty, t.entry_price, t.stop_price, t.risk_amount, t.total_commissions,
            t.status, t.realized_pnl, t.realized_R, t.schwab_transaction_id, t.notes,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER


def write_market_data(wb: Workbook, rows: Sequence[MarketDataRow]) -> None:
    ws = _get_or_create_sheet(wb, "Market_Data")
    if not rows:
        return

    # Collect all tickers that appear
    all_tickers = sorted({t for r in rows for t in r.prices})
    headers = ["Date"] + all_tickers
    _style_header(ws, headers)

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    for row_idx, r in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=r.date.isoformat()).font = BODY_FONT
        for col, ticker in enumerate(all_tickers, start=2):
            val = r.prices.get(ticker)
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = BODY_FONT
            cell.number_format = "0.00"
            cell.border = THIN_BORDER


def write_reference_data(wb: Workbook, rows: Sequence[ReferenceData]) -> None:
    ws = _get_or_create_sheet(wb, "Reference_Data")
    headers = ["Ticker", "Company", "Sector", "Industry", "Country", "Beta", "AssetClass"]
    _style_header(ws, headers)

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    for row_idx, r in enumerate(rows, start=2):
        values = [r.ticker, r.company, r.sector, r.industry, r.country, r.beta, r.asset_class]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER


def write_daily_units(wb: Workbook, rows: Sequence[DailyUnitsRow]) -> None:
    ws = _get_or_create_sheet(wb, "Daily_Units")
    if not rows:
        return

    all_tickers = sorted({t for r in rows for t in r.units})
    headers = ["Date"] + all_tickers
    _style_header(ws, headers)

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    for row_idx, r in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=r.date.isoformat()).font = BODY_FONT
        for col, ticker in enumerate(all_tickers, start=2):
            val = r.units.get(ticker, 0.0)
            cell = ws.cell(row=row_idx, column=col, value=val if abs(val) > 1e-9 else None)
            cell.font = BODY_FONT
            cell.number_format = "0.00"
            cell.border = THIN_BORDER


def write_daily_cash(wb: Workbook, rows: Sequence[DailyCashRow]) -> None:
    ws = _get_or_create_sheet(wb, "Daily_Cash")
    headers = ["Date", "Account", "Cash_Balance"]
    _style_header(ws, headers)

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    for row_idx, r in enumerate(rows, start=2):
        values = [r.date.isoformat(), r.account_id, r.cash_balance]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if col == 3:
                cell.number_format = "#,##0.00"


def write_daily_invested(wb: Workbook, rows: Sequence[DailyInvestedRow]) -> None:
    ws = _get_or_create_sheet(wb, "Daily_Invested")
    headers = ["Date", "Invested_Value"]
    _style_header(ws, headers)

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    for row_idx, r in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=r.date.isoformat()).font = BODY_FONT
        cell = ws.cell(row=row_idx, column=2, value=r.invested_value)
        cell.font = BODY_FONT
        cell.number_format = "#,##0.00"
        cell.border = THIN_BORDER


def write_cost_basis(wb: Workbook, rows: Sequence[CostBasisRow]) -> None:
    ws = _get_or_create_sheet(wb, "Cost_Basis")
    headers = ["Ticker", "Avg_Cost", "Current_Shares", "Total_Cost_Basis", "Account"]
    _style_header(ws, headers)

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    for row_idx, r in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=r.ticker).font = BODY_FONT
        c2 = ws.cell(row=row_idx, column=2, value=r.avg_cost)
        c2.font = BODY_FONT
        c2.number_format = "0.0000"
        c3 = ws.cell(row=row_idx, column=3, value=r.current_shares)
        c3.font = BODY_FONT
        c3.number_format = "#,##0.00"
        c4 = ws.cell(row=row_idx, column=4, value=r.total_cost_basis)
        c4.font = BODY_FONT
        c4.number_format = "$#,##0.00"
        ws.cell(row=row_idx, column=5, value=r.account_id).font = BODY_FONT
        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).border = THIN_BORDER


def write_data_lake(
    workbook_path: Path,
    trades: Sequence[TradeRecord],
    market_data: Sequence[MarketDataRow],
    reference_data: Sequence[ReferenceData],
    daily_units: Sequence[DailyUnitsRow],
    daily_cash: Sequence[DailyCashRow],
    daily_invested: Sequence[DailyInvestedRow] | None = None,
    cost_basis: Sequence[CostBasisRow] | None = None,
) -> Path:
    """
    Write (or overwrite) all Data Lake sheets and save the workbook.
    Creates the file if it does not exist.
    """
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb = _ensure_workbook(workbook_path)

    write_trade_log(wb, trades)
    write_market_data(wb, market_data)
    write_reference_data(wb, reference_data)
    write_daily_units(wb, daily_units)
    write_daily_cash(wb, daily_cash)
    if daily_invested is not None:
        write_daily_invested(wb, daily_invested)
    if cost_basis is not None:
        write_cost_basis(wb, cost_basis)

    for name in (
        "Dashboard", "Positions", "Equity_Curve", "Sector_Exposure",
        "Geographic_Exposure", "Correlation", "Risk_Metrics",
        "Stress_Testing", "Investment_Theses",
    ):
        _get_or_create_sheet(wb, name)

    wb.save(workbook_path)
    logger.info("Workbook saved → %s", workbook_path)
    return workbook_path
