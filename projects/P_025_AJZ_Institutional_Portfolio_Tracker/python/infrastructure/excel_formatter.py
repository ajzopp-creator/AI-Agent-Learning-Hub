"""
P_025 Infrastructure — Excel Formatter

openpyxl helpers for Bloomberg-style formatting and Analytics sheet setup.
Contains no business logic beyond presentation.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Colour & font constants (Bloomberg-inspired)
# ---------------------------------------------------------------------------
NAVY = "1C2541"
WHITE = "FFFFFF"
GREEN = "007A33"
RED = "B81D13"
LIGHT_GRAY = "F2F2F2"
MED_GRAY = "D0D0D0"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name="Segoe UI", bold=True, color=WHITE, size=11)
BODY_FONT = Font(name="Segoe UI", size=10)
TITLE_FONT = Font(name="Segoe UI", bold=True, size=14, color=NAVY)
KPI_LABEL_FONT = Font(name="Segoe UI", bold=True, size=10, color=NAVY)
KPI_VALUE_FONT = Font(name="Segoe UI", bold=True, size=12)

THIN_BORDER = Border(
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
    top=Side(style="thin", color=MED_GRAY),
    bottom=Side(style="thin", color=MED_GRAY),
)

POSITIVE_FONT = Font(name="Segoe UI", size=10, color=GREEN)
NEGATIVE_FONT = Font(name="Segoe UI", size=10, color=RED)


def style_header_row(ws: Worksheet, headers: list[str], row: int = 1) -> None:
    """Write and style a header row."""
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.freeze_panes = f"A{row + 1}"
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row}"


def auto_column_width(ws: Worksheet, min_width: int = 10, max_width: int = 22) -> None:
    """Set reasonable column widths based on header length."""
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        width = max(min_width, min(max_width, (len(str(header)) + 4) if header else min_width))
        ws.column_dimensions[get_column_letter(col)].width = width


def apply_number_formats(
    ws: Worksheet,
    start_row: int,
    end_row: int,
    col_formats: dict[int, str],
) -> None:
    """Apply number formats to a range of columns."""
    for row in range(start_row, end_row + 1):
        for col, fmt in col_formats.items():
            cell = ws.cell(row=row, column=col)
            cell.number_format = fmt
            cell.font = BODY_FONT
            cell.border = THIN_BORDER


def write_kpi_block(
    ws: Worksheet,
    start_row: int,
    start_col: int,
    labels: list[str],
) -> None:
    """Write a vertical KPI label + value block (values left blank for formulas)."""
    for i, label in enumerate(labels):
        r = start_row + i
        label_cell = ws.cell(row=r, column=start_col, value=label)
        label_cell.font = KPI_LABEL_FONT
        label_cell.alignment = Alignment(horizontal="right")
        value_cell = ws.cell(row=r, column=start_col + 1, value=None)
        value_cell.font = KPI_VALUE_FONT
        value_cell.border = THIN_BORDER


def clear_sheet(ws: Worksheet) -> None:
    """Remove all cells from a sheet while keeping the sheet object."""
    if ws.max_row > 0 and ws.max_column > 0:
        ws.delete_rows(1, ws.max_row)
