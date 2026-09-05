"""
P_025 Infrastructure — Sector & Correlation Analytics Sheets

Separated from analytics_sheets.py to stay under the 300-line hard limit.
"""

from __future__ import annotations

import logging

from openpyxl.workbook.workbook import Workbook

from domain.formula_templates import sector_exposure_headers
from infrastructure.excel_formatter import (
    BODY_FONT,
    HEADER_FILL,
    HEADER_FONT,
    THIN_BORDER,
    TITLE_FONT,
    auto_column_width,
    clear_sheet,
    style_header_row,
)

logger = logging.getLogger(__name__)


def build_sector_exposure(wb: Workbook) -> None:
    """Sector $ and % from Positions Market Value + Reference_Data Sector."""
    ws = wb["Sector_Exposure"]
    clear_sheet(ws)
    headers = sector_exposure_headers()
    style_header_row(ws, headers)

    ws.insert_rows(1)
    title = ws.cell(1, 1, "Sector Exposure — AJZ6348")
    title.font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(2, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    ws.freeze_panes = "A3"

    ref = wb["Reference_Data"]
    sectors: list[str] = []
    seen: set[str] = set()
    for row in range(2, ref.max_row + 1):
        s = ref.cell(row, 3).value
        if s and str(s).strip() and str(s).strip() not in seen:
            seen.add(str(s).strip())
            sectors.append(str(s).strip())
    sectors = sorted(sectors)[:40]

    for i, sector in enumerate(sectors):
        r = i + 3
        ws.cell(r, 1, sector).font = BODY_FONT
        formula_mv = (
            f'=IFERROR(SUMPRODUCT((Reference_Data!$C$2:$C$500=A{r})*'
            f'(Positions!$A$3:$A$202=Reference_Data!$A$2:$A$500)*'
            f'Positions!$D$3:$D$202),0)'
        )
        ws.cell(r, 2, formula_mv).font = BODY_FONT
        ws.cell(r, 2).number_format = "$#,##0.00"
        ws.cell(r, 3, f'=IF(SUM($B$3:$B$50)=0,0,B{r}/SUM($B$3:$B$50))').font = BODY_FONT
        ws.cell(r, 3).number_format = "0.00%"
        ws.cell(r, 4, f'=COUNTIF(Reference_Data!$C$2:$C$500,A{r})').font = BODY_FONT

        for col in range(1, 5):
            ws.cell(r, col).border = THIN_BORDER

    auto_column_width(ws)
    logger.info("Sector_Exposure sheet built with %d sectors", len(sectors))


def _correl_tickers(wb: Workbook, cap: int) -> list[str]:
    tickers: list[str] = []
    if "Positions" in wb.sheetnames:
        pos = wb["Positions"]
        for row in range(3, min(pos.max_row + 1, 3 + cap)):
            t = pos.cell(row, 1).value
            if t:
                tickers.append(str(t).strip().upper())
    if not tickers and "Reference_Data" in wb.sheetnames:
        ref = wb["Reference_Data"]
        for row in range(2, min(ref.max_row + 1, 1 + cap)):
            t = ref.cell(row, 1).value
            if t:
                tickers.append(str(t).strip().upper())
    return tickers[:cap]


def build_correlation(wb: Workbook) -> None:
    """CORREL of Market_Data close columns; diagonal = 1. Cap from config."""
    from config import CORREL_TICKER_CAP

    ws = wb["Correlation"]
    clear_sheet(ws)
    ws.cell(1, 1, "Correlation — Market_Data closes (top Positions)").font = TITLE_FONT
    tickers = _correl_tickers(wb, CORREL_TICKER_CAP)

    # Title row 1, header row 4, first data row 5 so (5,3) is off-diagonal CORREL.
    for i, t in enumerate(tickers):
        cell_h = ws.cell(4, i + 2, t)
        cell_h.fill = HEADER_FILL
        cell_h.font = HEADER_FONT
        cell_h.border = THIN_BORDER
        cell_v = ws.cell(i + 5, 1, t)
        cell_v.fill = HEADER_FILL
        cell_v.font = HEADER_FONT
        cell_v.border = THIN_BORDER

    for i in range(len(tickers)):
        for j in range(len(tickers)):
            cell = ws.cell(i + 5, j + 2)
            cell.border = THIN_BORDER
            cell.font = BODY_FONT
            cell.number_format = "0.00"
            if i == j:
                cell.value = 1
                continue
            header_cell = ws.cell(4, j + 2).coordinate
            row_ticker = ws.cell(i + 5, 1).coordinate
            cell.value = (
                f'=IFERROR(CORREL('
                f'INDEX(Market_Data!$B$2:$ZZ$4000,0,MATCH({row_ticker},Market_Data!$1:$1,0)),'
                f'INDEX(Market_Data!$B$2:$ZZ$4000,0,MATCH({header_cell},Market_Data!$1:$1,0))'
                f'),"")'
            )

    ws.column_dimensions["A"].width = 12
    logger.info("Correlation sheet filled with %d tickers", len(tickers))
