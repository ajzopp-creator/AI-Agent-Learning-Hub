"""Geographic exposure and stress-test sheets. I/O + formulas only."""

from __future__ import annotations

import logging

from openpyxl.workbook.workbook import Workbook

from infrastructure.excel_formatter import (
    BODY_FONT,
    HEADER_FILL,
    HEADER_FONT,
    THIN_BORDER,
    TITLE_FONT,
    auto_column_width,
    clear_sheet,
)

logger = logging.getLogger(__name__)


def build_geographic_exposure(wb: Workbook) -> None:
    """Country $ and % from Positions MV + Reference_Data Country (col E)."""
    ws = wb["Geographic_Exposure"]
    clear_sheet(ws)
    ws.cell(1, 1, "Geographic Exposure — PRIMARY accounts").font = TITLE_FONT
    headers = ["Country", "Market Value", "Weight %", "Position Count"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(2, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    ws.freeze_panes = "A3"

    countries: list[str] = []
    seen: set[str] = set()
    if "Reference_Data" in wb.sheetnames:
        ref = wb["Reference_Data"]
        for row in range(2, ref.max_row + 1):
            c = ref.cell(row, 5).value
            if c and str(c).strip() and str(c).strip() not in seen:
                seen.add(str(c).strip())
                countries.append(str(c).strip())
    countries = sorted(countries)[:40]

    for i, country in enumerate(countries):
        r = i + 3
        ws.cell(r, 1, country).font = BODY_FONT
        ws.cell(r, 2).font = BODY_FONT
        ws.cell(r, 2).number_format = "$#,##0.00"
        ws.cell(r, 2).value = (
            f'=IFERROR(SUMPRODUCT((Reference_Data!$E$2:$E$500=A{r})*'
            f'(Positions!$A$3:$A$202=Reference_Data!$A$2:$A$500)*'
            f'Positions!$D$3:$D$202),0)'
        )
        ws.cell(r, 3).value = f'=IF(SUM($B$3:$B$50)=0,0,B{r}/SUM($B$3:$B$50))'
        ws.cell(r, 3).number_format = "0.00%"
        ws.cell(r, 4).value = f'=COUNTIF(Reference_Data!$E$2:$E$500,A{r})'
        for col in range(1, 5):
            ws.cell(r, col).border = THIN_BORDER
            ws.cell(r, col).font = BODY_FONT

    auto_column_width(ws)
    logger.info("Geographic_Exposure built with %d countries", len(countries))


def build_stress_testing(wb: Workbook) -> None:
    """Linear MV shocks. Not options-aware. No rate series in v1."""
    ws = wb["Stress_Testing"]
    clear_sheet(ws)
    ws.cell(1, 1, "Stress Testing — linear MV shocks (not options-aware)").font = TITLE_FONT
    headers = ["Scenario", "Shock", "Stressed MV", "P&L vs Current MV", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(2, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    ws.freeze_panes = "A3"

    scenarios = [
        ("Base (current MV)", 0.0, "Unshocked Positions!D sum"),
        ("Equity -10%", -0.10, "Parallel equity shock"),
        ("Equity -20%", -0.20, "Parallel equity shock"),
        ("Equity -30%", -0.30, "Parallel equity shock"),
        ("Equity +10%", 0.10, "Parallel equity shock"),
        ("Rates +100 bp (label only)", 0.0, "No rate series in Data Lake — 0 shock"),
        ("Rates +200 bp (label only)", 0.0, "No rate series in Data Lake — 0 shock"),
    ]
    for i, (name, shock, note) in enumerate(scenarios):
        r = i + 3
        ws.cell(r, 1, name).font = BODY_FONT
        ws.cell(r, 2, shock).font = BODY_FONT
        ws.cell(r, 2).number_format = "0.00%"
        ws.cell(r, 3, f"=SUM(Positions!$D$3:$D$202)*(1+B{r})").font = BODY_FONT
        ws.cell(r, 3).number_format = "$#,##0.00"
        ws.cell(r, 4, f"=C{r}-SUM(Positions!$D$3:$D$202)").font = BODY_FONT
        ws.cell(r, 4).number_format = "$#,##0.00"
        ws.cell(r, 5, note).font = BODY_FONT
        for col in range(1, 6):
            ws.cell(r, col).border = THIN_BORDER

    auto_column_width(ws)
    logger.info("Stress_Testing built with %d scenarios", len(scenarios))
