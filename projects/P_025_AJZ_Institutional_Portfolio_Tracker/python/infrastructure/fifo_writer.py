"""Write Fifo_Lots and Fifo_Cost Data Lake sheets. I/O only."""

from __future__ import annotations

from typing import Sequence

from openpyxl.workbook.workbook import Workbook

from infrastructure.excel_writer import BODY_FONT, _get_or_create_sheet, _style_header
from schemas import FifoCostRow, FifoLotRow


def write_fifo_lots(wb: Workbook, rows: Sequence[FifoLotRow]) -> None:
    ws = _get_or_create_sheet(wb, "Fifo_Lots")
    headers = [
        "Account",
        "Ticker",
        "Open_Date",
        "Remaining_Qty",
        "Lot_Price",
        "Remaining_Cost",
        "Source_Trade_Id",
    ]
    _style_header(ws, headers)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    for i, r in enumerate(rows, start=2):
        values = [
            r.account_id,
            r.ticker,
            r.open_date.isoformat(),
            r.remaining_qty,
            r.lot_price,
            r.remaining_cost,
            r.source_trade_id,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(i, col, val)
            cell.font = BODY_FONT
            if col in (4, 5, 6):
                cell.number_format = "#,##0.00"


def write_fifo_cost(wb: Workbook, rows: Sequence[FifoCostRow]) -> None:
    ws = _get_or_create_sheet(wb, "Fifo_Cost")
    headers = ["Ticker", "Account", "Remaining_Shares", "Remaining_Cost"]
    _style_header(ws, headers)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    for i, r in enumerate(rows, start=2):
        ws.cell(i, 1, r.ticker).font = BODY_FONT
        ws.cell(i, 2, r.account_id).font = BODY_FONT
        c3 = ws.cell(i, 3, r.remaining_shares)
        c3.font = BODY_FONT
        c3.number_format = "#,##0.00"
        c4 = ws.cell(i, 4, r.remaining_cost)
        c4.font = BODY_FONT
        c4.number_format = "$#,##0.00"
