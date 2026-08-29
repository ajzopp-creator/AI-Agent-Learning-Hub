"""
P_025 Infrastructure — Analytics Sheet Builders

Builds Positions, Equity_Curve, Dashboard, and Risk_Metrics sheets.
openpyxl I/O only; formula patterns come from domain.formula_templates.
"""

from __future__ import annotations

import logging
from datetime import datetime

from openpyxl.workbook.workbook import Workbook

from domain.formula_templates import (
    dashboard_kpi_labels,
    equity_curve_headers,
    positions_headers,
    risk_metrics_labels,
)
from infrastructure.excel_formatter import (
    BODY_FONT,
    HEADER_FILL,
    HEADER_FONT,
    KPI_LABEL_FONT,
    KPI_VALUE_FONT,
    THIN_BORDER,
    TITLE_FONT,
    auto_column_width,
    clear_sheet,
    style_header_row,
    write_kpi_block,
)

logger = logging.getLogger(__name__)


def build_positions(wb: Workbook) -> None:
    """Populate the Positions sheet with headers + formulas."""
    ws = wb["Positions"]
    clear_sheet(ws)
    headers = positions_headers()
    style_header_row(ws, headers)

    ws.insert_rows(1)
    title = ws.cell(1, 1, "Positions — AJZ6348 (Primary View)")
    title.font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(2, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    ws.freeze_panes = "A3"

    ref = wb["Reference_Data"]
    tickers = []
    for row in range(2, ref.max_row + 1):
        t = ref.cell(row, 1).value
        if t:
            tickers.append(str(t).strip().upper())
    tickers = tickers[:200]

    for i, ticker in enumerate(tickers):
        r = i + 3
        ws.cell(r, 1, ticker).font = BODY_FONT
        shares_formula = (
            f'=IFERROR(LOOKUP(2,1/(Daily_Units!A:A<>""),'
            f'INDEX(Daily_Units!A:ZZ,0,MATCH(A{r},Daily_Units!$1:$1,0))),0)'
        )
        ws.cell(r, 2, shares_formula).font = BODY_FONT
        price_formula = (
            f'=IFERROR(LOOKUP(2,1/(Market_Data!A:A<>""),'
            f'INDEX(Market_Data!A:ZZ,0,MATCH(A{r},Market_Data!$1:$1,0))),0)'
        )
        ws.cell(r, 3, price_formula).font = BODY_FONT
        ws.cell(r, 4, f"=B{r}*C{r}").font = BODY_FONT
        # Cost Basis = Total_Cost_Basis from Cost_Basis helper (Python average-cost)
        cost_formula = (
            f'=IFERROR(VLOOKUP(A{r},Cost_Basis!$A:$D,4,FALSE),0)'
        )
        ws.cell(r, 5, cost_formula).font = BODY_FONT
        ws.cell(r, 6, f"=D{r}-E{r}").font = BODY_FONT
        ws.cell(r, 7, f'=IF(E{r}=0,0,F{r}/E{r})').font = BODY_FONT
        ws.cell(r, 8, f'=IF(SUM($D$3:$D$202)=0,0,D{r}/SUM($D$3:$D$202))').font = BODY_FONT
        ws.cell(r, 9, "AJZ6348").font = BODY_FONT

        for col in range(1, 10):
            ws.cell(r, col).border = THIN_BORDER

        ws.cell(r, 2).number_format = "#,##0.00"
        ws.cell(r, 3).number_format = "0.00"
        ws.cell(r, 4).number_format = "$#,##0.00"
        ws.cell(r, 5).number_format = "$#,##0.00"
        ws.cell(r, 6).number_format = "$#,##0.00"
        ws.cell(r, 7).number_format = "0.00%"
        ws.cell(r, 8).number_format = "0.00%"

    auto_column_width(ws)
    logger.info("Positions sheet built with %d tickers", len(tickers))


def build_equity_curve(wb: Workbook) -> None:
    """Populate Equity_Curve with date-driven formulas."""
    ws = wb["Equity_Curve"]
    clear_sheet(ws)
    headers = equity_curve_headers()
    style_header_row(ws, headers)

    cash = wb["Daily_Cash"]
    dates = []
    seen: set = set()
    for row in range(2, cash.max_row + 1):
        d = cash.cell(row, 1).value
        if d and d not in seen:
            seen.add(d)
            dates.append(d)
    dates = sorted(dates)

    for i, d in enumerate(dates):
        r = i + 2
        ws.cell(r, 1, d).font = BODY_FONT
        ws.cell(
            r, 2,
            f'=SUMIFS(Daily_Cash!C:C,Daily_Cash!A:A,A{r},Daily_Cash!B:B,"AJZ6348")',
        ).font = BODY_FONT
        # Invested Value from Daily_Invested helper series
        ws.cell(
            r, 3,
            f'=IFERROR(SUMIF(Daily_Invested!A:A,A{r},Daily_Invested!B:B),0)',
        ).font = BODY_FONT
        ws.cell(r, 4, f"=B{r}+C{r}").font = BODY_FONT
        if i == 0:
            ws.cell(r, 5, 0).font = BODY_FONT
            ws.cell(r, 6, 0).font = BODY_FONT
            ws.cell(r, 8, f"=D{r}").font = BODY_FONT
        else:
            ws.cell(r, 5, f'=IF(D{r-1}=0,0,(D{r}-D{r-1})/D{r-1})').font = BODY_FONT
            ws.cell(r, 6, f"=(1+F{r-1})*(1+E{r})-1").font = BODY_FONT
            ws.cell(r, 8, f"=MAX(H{r-1},D{r})").font = BODY_FONT
        ws.cell(r, 7, f'=IF(H{r}=0,0,(D{r}-H{r})/H{r})').font = BODY_FONT

        for col in range(1, 9):
            ws.cell(r, col).border = THIN_BORDER

        ws.cell(r, 2).number_format = "$#,##0.00"
        ws.cell(r, 3).number_format = "$#,##0.00"
        ws.cell(r, 4).number_format = "$#,##0.00"
        ws.cell(r, 5).number_format = "0.00%"
        ws.cell(r, 6).number_format = "0.00%"
        ws.cell(r, 7).number_format = "0.00%"
        ws.cell(r, 8).number_format = "$#,##0.00"

    auto_column_width(ws)
    logger.info("Equity_Curve sheet built with %d dates", len(dates))


def build_dashboard(wb: Workbook) -> None:
    """Populate Dashboard with KPI labels and simple linked formulas."""
    ws = wb["Dashboard"]
    clear_sheet(ws)

    ws.cell(1, 1, "AJZ Institutional Portfolio Tracker — Dashboard").font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws.cell(2, 1, "Primary Account: AJZ6348").font = KPI_LABEL_FONT
    ws.cell(3, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = BODY_FONT

    labels = dashboard_kpi_labels()
    write_kpi_block(ws, start_row=5, start_col=1, labels=labels)

    ws.cell(5, 2, '=IFERROR(LOOKUP(2,1/(Equity_Curve!D:D<>""),Equity_Curve!D:D),0)')
    ws.cell(5, 2).number_format = "$#,##0.00"
    ws.cell(6, 2, '=IFERROR(LOOKUP(2,1/(Equity_Curve!B:B<>""),Equity_Curve!B:B),0)')
    ws.cell(6, 2).number_format = "$#,##0.00"
    ws.cell(7, 2, "=B5-B6")
    ws.cell(7, 2).number_format = "$#,##0.00"
    ws.cell(8, 2, "=IFERROR(SUM(Positions!F:F),0)")
    ws.cell(8, 2).number_format = "$#,##0.00"
    ws.cell(9, 2, 0).number_format = "$#,##0.00"
    ws.cell(10, 2, 0).number_format = "0.00%"
    ws.cell(11, 2, '=IFERROR(LOOKUP(2,1/(Equity_Curve!F:F<>""),Equity_Curve!F:F),0)')
    ws.cell(11, 2).number_format = "0.00%"
    ws.cell(12, 2, "=IFERROR(MIN(Equity_Curve!G:G),0)")
    ws.cell(12, 2).number_format = "0.00%"
    ws.cell(13, 2, "=COUNTA(Positions!A:A)-2")
    ws.cell(14, 2, '=IFERROR(LOOKUP(2,1/(Equity_Curve!A:A<>""),Equity_Curve!A:A),"")')

    for r in range(5, 15):
        ws.cell(r, 2).font = KPI_VALUE_FONT
        ws.cell(r, 2).border = THIN_BORDER

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    logger.info("Dashboard sheet built")


def build_risk_metrics(wb: Workbook) -> None:
    """Populate Risk_Metrics with labels and core formulas from Equity_Curve."""
    ws = wb["Risk_Metrics"]
    clear_sheet(ws)

    ws.cell(1, 1, "Risk Metrics — AJZ6348").font = TITLE_FONT
    labels = risk_metrics_labels()
    write_kpi_block(ws, start_row=3, start_col=1, labels=labels)

    # Observation days (row 11)
    ws.cell(11, 2, "=COUNTA(Equity_Curve!E:E)-1")
    # Max Drawdown (row 7)
    ws.cell(7, 2, "=IFERROR(MIN(Equity_Curve!G:G),0)")
    ws.cell(7, 2).number_format = "0.00%"
    # Annualized Return ≈ last cumulative return annualized (row 3)
    # Using (1+cum)^(252/n)-1 approximation via Equity_Curve
    ws.cell(
        3, 2,
        '=IFERROR((1+LOOKUP(2,1/(Equity_Curve!F:F<>""),Equity_Curve!F:F))'
        '^(252/MAX(B11,1))-1,0)',
    )
    ws.cell(3, 2).number_format = "0.00%"
    # Annualized Volatility = STDEV of daily returns * SQRT(252) (row 4)
    ws.cell(4, 2, "=IFERROR(STDEV(Equity_Curve!E:E)*SQRT(252),0)")
    ws.cell(4, 2).number_format = "0.00%"
    # Sharpe = (AnnReturn - Rf) / AnnVol ; Rf approx 4.5% (row 5)
    ws.cell(5, 2, "=IFERROR((B3-0.045)/B4,0)")
    ws.cell(5, 2).number_format = "0.00"
    # Sortino / VaR / CVaR / Beta — still TBD (rows 6, 8, 9, 10)
    for r in [6, 8, 9, 10]:
        ws.cell(r, 2, "TBD")
        ws.cell(r, 2).font = BODY_FONT

    for r in range(3, 12):
        ws.cell(r, 2).border = THIN_BORDER
        if r not in (6, 8, 9, 10):
            ws.cell(r, 2).font = KPI_VALUE_FONT

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    logger.info("Risk_Metrics sheet built")


def apply_data_lake_formatting(wb: Workbook) -> None:
    """Light formatting pass on existing Data Lake sheets."""
    for name in (
        "Trade_Log", "Market_Data", "Reference_Data",
        "Daily_Units", "Daily_Cash", "Daily_Invested", "Cost_Basis",
    ):
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.freeze_panes = "A2"
            try:
                ws.auto_filter.ref = ws.dimensions
            except Exception:
                pass
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(1, col)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
    logger.info("Data Lake formatting applied")
