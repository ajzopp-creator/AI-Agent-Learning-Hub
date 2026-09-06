"""Tracker Dashboard reader — loads Tracker Excel into a TrackerLookup schema."""

import logging
from pathlib import Path
from typing import Optional

import openpyxl
from pydantic import ValidationError

from config import TRACKER_DASHBOARD
from schemas import TrackerEntry, TrackerLookup

logger = logging.getLogger(__name__)

# Signal source normalization - maps Tracker variants to valid system_ids
_SIGNAL_MAP = {
    "P_118_EDDIEZZ"              : "P_118",
    "P_118_EDDIEZ"               : "P_118",
    "EDDIEZ_P115RECHECK"         : "P_118",
    "P_910_COMBINED"             : "P_910",
    "P_910_V4"                   : "P_910",
    "P_910_RELATIVESTRENGTH"     : "P_910",
    "P_117_P910"                 : "P_117",
    "P_117_D130"                 : "P_117",
    "P_117_RELATIVESTRENGTH"     : "P_117",
    "P_117_D050"                 : "P_117",
    "P_117_GEMINI"               : "P_117",
    "P_117_P190"                 : "P_117",
    "P117:WALLSTZEN"             : "P_117",
}

_VALID_SYSTEMS = {"P_115","P_116","P_117","P_118","P_300","P_910","P_920","Day","SNT","TOS_Import","P_105","P_110","P_120","P_210"}

def _normalize_signal(raw: str) -> str:
    """Map Tracker signal_source variants to a valid system_id."""
    clean = raw.strip().upper()
    if clean in _SIGNAL_MAP:
        return _SIGNAL_MAP[clean]
    # If it already matches a valid system (case-insensitive), return it as-is
    for v in _VALID_SYSTEMS:
        if clean == v.upper():
            return v
    return "TOS_Import"



# ── Column detection ───────────────────────────────────────────────────────

def _find_columns(sheet) -> dict:
    """Locate required columns from the Tracker Dashboard header row.

    Normalizes header names by stripping spaces and lowercasing so both
    'SignalSource' and 'Signal Source' resolve correctly.

    Args:
        sheet: Active openpyxl worksheet.

    Returns:
        Dict mapping field name → 1-based column index.

    Raises:
        ValueError: If any required column cannot be found.
    """
    raw_header = [
        str(cell.value).strip() if cell.value else ""
        for cell in next(sheet.iter_rows(min_row=1, max_row=1))
    ]
    # Normalized (no spaces, lowercase) → original index
    normalized = {h.lower().replace(" ", ""): i + 1 for i, h in enumerate(raw_header)}

    # Required mappings: field_name → possible normalized header names
    required = {
        "trade_date"   : ["date", "tradedate", "buydate", "opendate"],
        "symbol"       : ["symbol", "buy", "ticker"],
        "signal_source": ["signalsource", "signal", "source", "system"],
        "traded"       : ["traded", "trade"],
    }
    # Optional mappings
    optional = {
        "entry_price": ["entryprice", "entry"],
        "outcome"    : ["outcome", "result"],
        "sl_level"   : ["sllevel", "sl_level", "sl"],
        "stop_level" : ["stoplevel", "stop_level", "stop"],
    }

    col_map = {}
    for field, candidates in required.items():
        match = next((normalized[c] for c in candidates if c in normalized), None)
        if match is None:
            raise ValueError(
                f"Tracker Dashboard missing required column for '{field}'. "
                f"Tried: {candidates}. Found headers: {list(normalized.keys())}"
            )
        col_map[field] = match

    for field, candidates in optional.items():
        match = next((normalized[c] for c in candidates if c in normalized), None)
        if match:
            col_map[field] = match

    logger.debug(f"Tracker column map: {col_map}")
    return col_map


# ── Main loader ────────────────────────────────────────────────────────────

def load_tracker_lookup(
    tracker_path: Optional[Path] = None,
    require_traded: bool = False,
) -> Optional[TrackerLookup]:
    """Read Tracker Dashboard and return a validated TrackerLookup object.

    Args:
        tracker_path: Path to Tracker Dashboard xlsx. Defaults to config value.
        require_traded: If True (default), only rows where Traded='Yes' are
            loaded. Pass False for backfill/resolution tasks where any row
            with a SignalSource is useful regardless of traded status.

    Returns:
        Populated TrackerLookup object, or None on failure.
    """
    path = tracker_path or TRACKER_DASHBOARD

    if not path.exists():
        logger.warning(
            f"Tracker Dashboard not found: {path} — defaulting to TOS_Import."
        )
        return None

    try:
        wb    = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        sheet = wb.active
        col_map = _find_columns(sheet)

        lookup = TrackerLookup(source_file=path.name)
        skipped = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            lookup.total_rows += 1

            # Extract raw values using column map
            raw = {
                field: row[col_idx - 1]
                for field, col_idx in col_map.items()
                if col_idx - 1 < len(row)
            }

            # Skip rows with missing required fields
            if not raw.get("trade_date") or not raw.get("symbol") or not raw.get("signal_source"):
                skipped += 1
                continue

            # Skip rows where signal_source is a placeholder
            sig = str(raw.get("signal_source", "")).strip()
            if sig in ("-", "", "None", "N/A"):
                skipped += 1
                continue

            try:
                entry = TrackerEntry(**raw)
            except ValidationError as e:
                logger.debug(f"Skipping invalid Tracker row: {raw} — {e}")
                skipped += 1
                continue

            if require_traded and not entry.traded:
                skipped += 1
                continue

            lookup.entries[entry.lookup_key] = _normalize_signal(entry.signal_source)

            # Resolve stop price: StopLevel first, SLLevel fallback, None if both absent
            stop_val = entry.stop_level if entry.stop_level is not None else entry.sl_level
            lookup.stop_prices[entry.lookup_key] = stop_val

            lookup.traded_rows += 1

        lookup.skipped_rows = skipped
        wb.close()

        logger.info(lookup.summary())
        return lookup

    except PermissionError:
        logger.warning(
            "Tracker Dashboard is open in Excel — close it and retry. "
            "Defaulting to TOS_Import for this run."
        )
        return None
    except ValueError as e:
        logger.warning(f"Tracker Dashboard column error: {e} — defaulting to TOS_Import.")
        return None
    except Exception as e:
        logger.warning(f"Tracker Dashboard read error: {e} — defaulting to TOS_Import.")
        return None


def match_system(
    lookup: Optional[TrackerLookup],
    symbol: str,
    open_date: str,
    default: str = "TOS_Import",
) -> str:
    """Look up the system name for a symbol + date pair.

    Args:
        lookup: TrackerLookup from load_tracker_lookup(), or None.
        symbol: Underlying symbol to match (normalized internally).
        open_date: Trade open date as 'YYYY-MM-DD' string.
        default: Fallback system name if no match found.

    Returns:
        Matched system_id string, or default if not found.
    """
    if lookup is None:
        return default
    result = lookup.get(symbol, open_date, default)
    if result == default:
        logger.debug(f"No Tracker match for ({symbol}, {open_date}) — using '{default}'.")
    return result

def match_stop_price(lookup, symbol: str, open_date: str):
    """Return stop price from Tracker for a paper trade, or None if unavailable.

    Only called for PAPER account trades.

    Args:
        lookup: TrackerLookup object or None.
        symbol: Underlying symbol.
        open_date: Trade open date as YYYY-MM-DD string.

    Returns:
        Stop price as float, or None if not found.
    """
    if lookup is None:
        return None
    return lookup.get_stop(symbol, open_date)
