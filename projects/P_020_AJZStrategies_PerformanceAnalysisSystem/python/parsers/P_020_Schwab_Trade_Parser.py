"""
P_020 Schwab Trade Parser -- Phase 2C
Reads raw JSON pull file, produces OPTIONS and STOCKS import CSVs
matching Excel log column structure.

Usage:
    python P_020_Schwab_Trade_Parser.py                          # latest file in api_pulls\ajz_strategies\
    python P_020_Schwab_Trade_Parser.py --file path\to\file.json # specific file
    python P_020_Schwab_Trade_Parser.py --account IRA            # latest IRA pull
"""

import json
import csv
import sys
import argparse
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict

# -- Paths --------------------------------------------------------------------
BASE_DIR      = Path(__file__).resolve().parent.parent.parent
PULL_BASE_DIR = BASE_DIR / "data" / "api_pulls"
OUTPUT_DIR    = BASE_DIR / "data" / "processed"
TRACKER_PATH  = Path(r"C:\Users\Trader\Documents\AJZStrategiesLLC\P_115_TrackerAudit\P_115_118_TrackerDashboard_V2.xlsx")
AUDIT_DIR     = BASE_DIR / "audit_logs"

# -- Asset types to process ---------------------------------------------------
VALID_ASSET_TYPES = {"OPTION", "EQUITY"}

# -- Excel output columns -----------------------------------------------------
OPTIONS_COLUMNS = [
    "Symbol", "System", "Trade Type", "Long", "Trade Date", "Entry Price",
    "Contracts", "Exit #1", "# Exited", "Exit Date", "# of Days",
    "Exit #2", "# Exited2", "Exit Date3", "# of Days4",
    "Comm.", "Gain/Loss", "Trade Comments", "Exit #1 Gain", "Exit #2 Gain"
]

STOCKS_COLUMNS = [
    "Symbol", "System", "Long/Short", "Trade Date", "Entry Price", "Shares",
    "Exit #1", "# Exited", "Exit Date", "# of Days",
    "Exit #2", "# Exited2", "Exit Date3", "# of Days4",
    "Comm.", "Gain/Loss", "ROI", "Trade Comments",
    "Exit #1 Gain", "Exit #2 Gain", "Total Gain", "Total ROI %",
    "R:R Ratio", "Win/Loss", "Strategy Notes", "Review Status"
]

# -- Tracker Dashboard lookup -------------------------------------------------
def load_tracker(tracker_path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(tracker_path, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows())]

        # Find symbol, date, system columns
        sym_col = date_col = sys_col = None
        for i, h in enumerate(headers):
            hl = h.lower()
            if hl in ("symbol", "buy"):
                sym_col = i
            if hl in ("date", "buy date", "trade date"):
                date_col = i
            if hl in ("signal source", "system", "signalsource"):
                sys_col = i

        if sym_col is None or date_col is None or sys_col is None:
            print(f"  WARNING: Tracker column mapping failed. Headers: {headers}", flush=True)
            return {}

        lookup = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            sym  = str(row[sym_col]).strip().upper() if row[sym_col] else None
            dt   = row[date_col]
            sys_ = str(row[sys_col]).strip() if row[sys_col] else None
            if sym and dt and sys_:
                if isinstance(dt, datetime):
                    date_str = dt.strftime("%Y-%m-%d")
                else:
                    try:
                        date_str = datetime.strptime(str(dt), "%m/%d/%Y").strftime("%Y-%m-%d")
                    except:
                        date_str = str(dt)
                lookup[(sym, date_str)] = sys_
        print(f"  Tracker loaded: {len(lookup)} entries", flush=True)
        return lookup

    except FileNotFoundError:
        print(f"  WARNING: Tracker not found at {tracker_path} -- defaulting to TOS_Import", flush=True)
        return {}
    except Exception as e:
        print(f"  WARNING: Tracker load error: {e} -- defaulting to TOS_Import", flush=True)
        return {}

def match_system(symbol, trade_date, tracker, asset_type="EQUITY"):
    if not tracker:
        return "TOS_Import"
    sym = symbol.upper()

    if asset_type == "OPTION":
        # ±3 day window — options signal dates may not match trade date exactly
        if isinstance(trade_date, datetime):
            base = trade_date
        else:
            base = datetime.strptime(str(trade_date), "%Y-%m-%d")
        for delta in range(0, 4):
            for sign in ([0] if delta == 0 else [1, -1]):
                check = (base + timedelta(days=delta * sign)).strftime("%Y-%m-%d")
                result = tracker.get((sym, check))
                if result:
                    return result
        return "TOS_Import"
    else:
        # Stocks: exact date match
        date_str = trade_date.strftime("%Y-%m-%d") if isinstance(trade_date, datetime) else str(trade_date)
        return tracker.get((sym, date_str), "TOS_Import")

# -- JSON helpers -------------------------------------------------------------
def parse_datetime(dt_str):
    for fmt in ("%Y-%m-%dT%H:%M:%S+0000", "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S%z"):
        try:
            dt = datetime.strptime(dt_str[:19], fmt[:len(fmt.split("+")[0])])
            return dt
        except:
            pass
    try:
        return datetime.fromisoformat(dt_str[:19])
    except:
        return None

def extract_trade_item(transfer_items):
    for item in transfer_items:
        if item.get("instrument", {}).get("assetType") in VALID_ASSET_TYPES:
            return item
    return None

def extract_fees(transfer_items):
    total = 0.0
    for item in transfer_items:
        if item.get("instrument", {}).get("assetType") == "CURRENCY":
            total += abs(item.get("amount", 0.0))
    return round(total, 2)

# -- Step 1: Parse raw transactions into normalized orders --------------------
def parse_transactions(transactions):
    """
    Group by orderId, aggregate fills.
    Returns list of order dicts.
    """
    order_groups = defaultdict(list)
    skipped = []

    for t in transactions:
        order_id = t.get("orderId")
        if not order_id:
            skipped.append({"reason": "no_order_id", "activityId": t.get("activityId")})
            continue

        trade_item = extract_trade_item(t.get("transferItems", []))
        if not trade_item:
            skipped.append({"reason": "no_valid_trade_item", "activityId": t.get("activityId")})
            continue

        inst = trade_item.get("instrument", {})
        asset_type = inst.get("assetType")

        order_groups[order_id].append({
            "activityId"    : t.get("activityId"),
            "orderId"       : order_id,
            "tradeDate"     : parse_datetime(t.get("tradeDate", "")),
            "assetType"     : asset_type,
            "symbol"        : inst.get("underlyingSymbol") if asset_type == "OPTION" else inst.get("symbol"),
            "putCall"       : inst.get("putCall", ""),
            "price"         : trade_item.get("price", 0.0),
            "amount"        : trade_item.get("amount", 0.0),
            "positionEffect": trade_item.get("positionEffect", ""),
            "fees"          : extract_fees(t.get("transferItems", [])),
        })

    # Aggregate fills within same orderId
    orders = []
    for order_id, fills in order_groups.items():
        if not fills:
            continue

        first = fills[0]
        total_amount = sum(abs(f["amount"]) for f in fills)
        total_fees   = round(sum(f["fees"] for f in fills), 2)

        # Weighted average price
        total_cost   = sum(abs(f["amount"]) * f["price"] for f in fills)
        avg_price    = round(total_cost / total_amount, 4) if total_amount else 0.0

        orders.append({
            "orderId"       : order_id,
            "tradeDate"     : first["tradeDate"],
            "assetType"     : first["assetType"],
            "symbol"        : first["symbol"],
            "putCall"       : first["putCall"],
            "price"         : avg_price,
            "quantity"      : total_amount,
            "positionEffect": first["positionEffect"],
            "fees"          : total_fees,
            "fillCount"     : len(fills),
        })

    orders.sort(key=lambda x: x["tradeDate"] or datetime.min)
    return orders, skipped

# -- Step 2: 10-minute consolidation for OPENING orders ----------------------
def consolidate_openings(orders):
    """
    Same symbol + OPENING + within 10 minutes → merge into one entry.
    """
    openings = [o for o in orders if o["positionEffect"] == "OPENING"]
    closings = [o for o in orders if o["positionEffect"] == "CLOSING"]

    consolidated = []
    used = set()

    for i, order in enumerate(openings):
        if i in used:
            continue
        group = [order]
        used.add(i)
        for j, other in enumerate(openings):
            if j in used:
                continue
            if other["symbol"] == order["symbol"] and other["assetType"] == order["assetType"]:
                delta = abs((other["tradeDate"] - order["tradeDate"]).total_seconds())
                if delta <= 600:  # 10 minutes
                    group.append(other)
                    used.add(j)

        if len(group) == 1:
            consolidated.append(order)
        else:
            total_qty  = sum(g["quantity"] for g in group)
            total_fees = round(sum(g["fees"] for g in group), 2)
            total_cost = sum(g["quantity"] * g["price"] for g in group)
            avg_price  = round(total_cost / total_qty, 4) if total_qty else 0.0
            consolidated.append({
                "orderId"       : group[0]["orderId"],
                "tradeDate"     : group[0]["tradeDate"],
                "assetType"     : group[0]["assetType"],
                "symbol"        : group[0]["symbol"],
                "putCall"       : group[0]["putCall"],
                "price"         : avg_price,
                "quantity"      : total_qty,
                "positionEffect": "OPENING",
                "fees"          : total_fees,
                "fillCount"     : sum(g["fillCount"] for g in group),
                "consolidated"  : len(group),
            })

    return consolidated, closings

# -- Step 3: Match entries to exits ------------------------------------------
def match_trades(openings, closings):
    """
    Match OPENING orders to CLOSING orders by symbol in date order.
    Returns matched positions and orphaned closings.
    """
    positions   = []
    orphaned    = []
    entry_pool  = defaultdict(list)

    for o in openings:
        entry_pool[o["symbol"]].append(o)

    for sym in entry_pool:
        entry_pool[sym].sort(key=lambda x: x["tradeDate"])

    sym_closings = defaultdict(list)
    for c in closings:
        sym_closings[c["symbol"]].append(c)
    for sym in sym_closings:
        sym_closings[sym].sort(key=lambda x: x["tradeDate"])

    # Match each opening to up to 2 closings
    for sym, entries in entry_pool.items():
        exits = sym_closings.get(sym, [])
        exit_idx = 0

        for entry in entries:
            pos = {
                "symbol"        : sym,
                "assetType"     : entry["assetType"],
                "putCall"       : entry["putCall"],
                "tradeDate"     : entry["tradeDate"],
                "entryPrice"    : entry["price"],
                "quantity"      : entry["quantity"],
                "entryFees"     : entry["fees"],
                "exit1"         : None,
                "exit1Qty"      : None,
                "exit1Date"     : None,
                "exit1Fees"     : 0.0,
                "exit2"         : None,
                "exit2Qty"      : None,
                "exit2Date"     : None,
                "exit2Fees"     : 0.0,
            }

            # Assign up to 2 exits
            exits_assigned = 0
            while exit_idx < len(exits) and exits_assigned < 2:
                ex = exits[exit_idx]
                if ex["tradeDate"] >= entry["tradeDate"]:
                    if exits_assigned == 0:
                        pos["exit1"]     = ex["price"]
                        pos["exit1Qty"]  = ex["quantity"]
                        pos["exit1Date"] = ex["tradeDate"]
                        pos["exit1Fees"] = ex["fees"]
                    else:
                        pos["exit2"]     = ex["price"]
                        pos["exit2Qty"]  = ex["quantity"]
                        pos["exit2Date"] = ex["tradeDate"]
                        pos["exit2Fees"] = ex["fees"]
                    exit_idx += 1
                    exits_assigned += 1
                else:
                    break

            positions.append(pos)

    # Any unmatched closings are orphaned
    all_matched_syms = set(entry_pool.keys())
    for sym, exits in sym_closings.items():
        if sym not in all_matched_syms:
            for ex in exits:
                orphaned.append(ex)
        else:
            # Check if there were more exits than entries could absorb
            entry_count = len(entry_pool[sym])
            assigned    = entry_count * 2
            if len(exits) > assigned:
                for ex in exits[assigned:]:
                    orphaned.append(ex)

    return positions, orphaned

# -- Step 4: Format output rows -----------------------------------------------
def days_between(d1, d2):
    if d1 and d2:
        return abs((d2 - d1).days)
    return ""

def fmt_date(dt):
    return dt.strftime("%#m/%#d/%y") if dt else ""

def build_options_row(pos, tracker):
    system = match_system(pos["symbol"], pos["tradeDate"], tracker, asset_type="OPTION")
    total_comm = round(pos["entryFees"] + pos["exit1Fees"] + pos["exit2Fees"], 2)

    return {
        "Symbol"        : pos["symbol"],
        "System"        : system,
        "Trade Type"    : pos["putCall"],
        "Long"          : "Long",
        "Trade Date"    : fmt_date(pos["tradeDate"]),
        "Entry Price"   : pos["entryPrice"],
        "Contracts"     : pos["quantity"],
        "Exit #1"       : pos["exit1"] or "",
        "# Exited"      : pos["exit1Qty"] or "",
        "Exit Date"     : fmt_date(pos["exit1Date"]),
        "# of Days"     : days_between(pos["tradeDate"], pos["exit1Date"]),
        "Exit #2"       : pos["exit2"] or "",
        "# Exited2"     : pos["exit2Qty"] or "",
        "Exit Date3"    : fmt_date(pos["exit2Date"]),
        "# of Days4"    : days_between(pos["tradeDate"], pos["exit2Date"]),
        "Comm."         : total_comm,
        "Gain/Loss"     : "",
        "Trade Comments": "",
        "Exit #1 Gain"  : "",
        "Exit #2 Gain"  : "",
    }

def build_stocks_row(pos, tracker):
    system = match_system(pos["symbol"], pos["tradeDate"], tracker)
    total_comm = round(pos["entryFees"] + pos["exit1Fees"] + pos["exit2Fees"], 2)

    return {
        "Symbol"        : pos["symbol"],
        "System"        : system,
        "Long/Short"    : "Long",
        "Trade Date"    : fmt_date(pos["tradeDate"]),
        "Entry Price"   : pos["entryPrice"],
        "Shares"        : pos["quantity"],
        "Exit #1"       : pos["exit1"] or "",
        "# Exited"      : pos["exit1Qty"] or "",
        "Exit Date"     : fmt_date(pos["exit1Date"]),
        "# of Days"     : days_between(pos["tradeDate"], pos["exit1Date"]),
        "Exit #2"       : pos["exit2"] or "",
        "# Exited2"     : pos["exit2Qty"] or "",
        "Exit Date3"    : fmt_date(pos["exit2Date"]),
        "# of Days4"    : days_between(pos["tradeDate"], pos["exit2Date"]),
        "Comm."         : total_comm,
        "Gain/Loss"     : "",
        "ROI"           : "",
        "Trade Comments": "",
        "Exit #1 Gain"  : "",
        "Exit #2 Gain"  : "",
        "Total Gain"    : "",
        "Total ROI %"   : "",
        "R:R Ratio"     : "",
        "Win/Loss"      : "",
        "Strategy Notes": "",
        "Review Status" : "",
    }

# -- Step 5: Write CSVs -------------------------------------------------------
def write_csv(rows, columns, out_path):
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)

# -- Step 6: Write audit log --------------------------------------------------
def write_audit(audit_lines, out_path):
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(audit_lines))
    print(f"\n  Audit log: {out_path.name}", flush=True)

# -- File selection -----------------------------------------------------------
def find_latest_pull(account_label):
    pull_dir = PULL_BASE_DIR / account_label.lower()
    if not pull_dir.exists():
        print(f"ERROR: Pull directory not found: {pull_dir}", flush=True)
        sys.exit(1)
    files = sorted(pull_dir.glob("P_020_raw_*.json"), key=lambda x: x.stat().st_mtime, reverse=True)
    if not files:
        print(f"ERROR: No pull files found in {pull_dir}", flush=True)
        sys.exit(1)
    return files[0]

# -- Main ---------------------------------------------------------------------
def run():
    parser = argparse.ArgumentParser(description="P_020 Schwab Trade Parser")
    parser.add_argument("--file",    default=None, help="Path to specific JSON pull file")
    parser.add_argument("--account", default="AJZ", help="Account label for auto-selecting latest file (default: AJZ)")
    args = parser.parse_args()

    # Resolve input file
    if args.file:
        input_path = Path(args.file)
    else:
        account_map = {"AJZ": "AJZ_Strategies", "IRA": "Inherited_Roth"}
        label = account_map.get(args.account.upper(), args.account)
        input_path = find_latest_pull(label)

    print(f"\nP_020 Schwab Trade Parser -- Phase 2C", flush=True)
    print("=" * 50, flush=True)
    print(f"Input file : {input_path.name}", flush=True)

    # Load JSON
    with open(input_path) as f:
        data = json.load(f)

    transactions   = data.get("transactions", [])
    account_label  = data.get("account_label", "Unknown")
    start_date     = data.get("start_date", "")
    end_date       = data.get("end_date", "")

    print(f"Account    : {account_label}", flush=True)
    print(f"Date range : {start_date} to {end_date}", flush=True)
    print(f"Transactions in file: {len(transactions)}", flush=True)

    # Load tracker
    print("\nLoading Tracker Dashboard...", flush=True)
    tracker = load_tracker(TRACKER_PATH)

    # Parse
    print("\nParsing transactions...", flush=True)
    orders, skipped = parse_transactions(transactions)
    print(f"  Orders after fill aggregation: {len(orders)}", flush=True)
    print(f"  Skipped (no orderId or no trade item): {len(skipped)}", flush=True)

    # Separate options and stocks
    option_orders = [o for o in orders if o["assetType"] == "OPTION"]
    stock_orders  = [o for o in orders if o["assetType"] == "EQUITY"]
    print(f"  Option orders: {len(option_orders)}", flush=True)
    print(f"  Stock orders : {len(stock_orders)}", flush=True)

    # Consolidate 10-min openings
    print("\nApplying 10-minute consolidation...", flush=True)
    opt_openings,  opt_closings  = consolidate_openings(option_orders)
    stk_openings,  stk_closings  = consolidate_openings(stock_orders)
    print(f"  Option openings after consolidation: {len(opt_openings)}", flush=True)
    print(f"  Stock openings after consolidation : {len(stk_openings)}", flush=True)

    # Match entries to exits
    print("\nMatching entries to exits...", flush=True)
    opt_positions, opt_orphans = match_trades(opt_openings, opt_closings)
    stk_positions, stk_orphans = match_trades(stk_openings, stk_closings)
    print(f"  Option positions matched: {len(opt_positions)}", flush=True)
    print(f"  Option orphaned exits   : {len(opt_orphans)}", flush=True)
    print(f"  Stock positions matched : {len(stk_positions)}", flush=True)
    print(f"  Stock orphaned exits    : {len(stk_orphans)}", flush=True)

    # Build output rows
    opt_rows = [build_options_row(p, tracker) for p in opt_positions]
    stk_rows = [build_stocks_row(p, tracker)  for p in stk_positions]

    # Write CSVs
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem        = f"{account_label}_{start_date}_to_{end_date}_{timestamp}"
    # Map account label to live/paper output folder
    account_upper = account_label.upper()
    if "IRA" in account_upper or "ROTH" in account_upper or "INHERITED" in account_upper:
        out_subdir = "paper"
    else:
        out_subdir = "live"
    out_dir = OUTPUT_DIR / out_subdir
    out_dir.mkdir(parents=True, exist_ok=True)

    opt_path = out_dir / f"P_020_{stem}_OPTIONS_IMPORT.csv"
    stk_path = out_dir / f"P_020_{stem}_STOCKS_IMPORT.csv"

    write_csv(opt_rows, OPTIONS_COLUMNS, opt_path)
    write_csv(stk_rows, STOCKS_COLUMNS,  stk_path)

    print(f"\n  Options CSV : {opt_path.name} ({len(opt_rows)} rows)", flush=True)
    print(f"  Stocks CSV  : {stk_path.name} ({len(stk_rows)} rows)", flush=True)

    # Audit log
    audit_lines = [
        f"P_020 Trade Parser Audit Log",
        f"Run timestamp : {datetime.now().isoformat()}",
        f"Input file    : {input_path.name}",
        f"Account       : {account_label}",
        f"Date range    : {start_date} to {end_date}",
        f"",
        f"--- PROCESSING SUMMARY ---",
        f"Total transactions in file : {len(transactions)}",
        f"Skipped (no orderId/item)  : {len(skipped)}",
        f"Orders after fill aggregation: {len(orders)}",
        f"  Option orders            : {len(option_orders)}",
        f"  Stock orders             : {len(stock_orders)}",
        f"",
        f"--- CONSOLIDATION ---",
        f"Option openings after 10-min consolidation: {len(opt_openings)}",
        f"Stock openings after 10-min consolidation : {len(stk_openings)}",
        f"",
        f"--- MATCHING ---",
        f"Option positions matched   : {len(opt_positions)}",
        f"Stock positions matched    : {len(stk_positions)}",
        f"",
        f"--- ORPHANED EXITS (no matching entry in date range) ---",
        f"Option orphans: {len(opt_orphans)}",
    ]
    for o in opt_orphans:
        audit_lines.append(f"  {o['symbol']} on {fmt_date(o['tradeDate'])} -- check prior period")

    audit_lines.append(f"Stock orphans : {len(stk_orphans)}")
    for o in stk_orphans:
        audit_lines.append(f"  {o['symbol']} on {fmt_date(o['tradeDate'])} -- check prior period")

    audit_lines += [
        f"",
        f"--- OUTPUT FILES ---",
        f"Options CSV : {opt_path.name}",
        f"Stocks CSV  : {stk_path.name}",
    ]

    audit_path = AUDIT_DIR / f"P_020_Parser_Audit_{timestamp}.txt"
    write_audit(audit_lines, audit_path)

    print("\n" + "=" * 50, flush=True)
    print("SUMMARY", flush=True)
    print("=" * 50, flush=True)
    print(f"  Options import rows : {len(opt_rows)}", flush=True)
    print(f"  Stocks import rows  : {len(stk_rows)}", flush=True)
    print(f"  Orphaned exits      : {len(opt_orphans) + len(stk_orphans)} (see audit log)", flush=True)
    print(f"\nPhase 2C complete.", flush=True)

if __name__ == "__main__":
    run()
