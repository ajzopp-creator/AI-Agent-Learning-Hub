"""
P_020 Trade Import Utility - Enhanced Version
==============================================
Imports trade data from TOS export CSV files to Excel log workbooks.

Features:
- Automatically matches System names from Tracker Dashboard (REQ-020126_01)
- Preserves formulas in calculated columns (REQ-020207_01)
- Handles both paper and live accounts
- Menu-driven interface

Requirements:
    pip install openpyxl pandas

Author: Anthony (AJZ Strategies LLC)
Version: 2.2
Date: 2026-02-07
"""

import pandas as pd
import openpyxl
import os
import json
from datetime import datetime
import re

# ============================================================
# CONFIGURATION
# ============================================================

# Base paths
AI_HUB_BASE = r"C:\Users\Trader\AI-Agent-Learning-Hub\projects\P_020_AJZStrategies_PerformanceAnalysisSystem"
LIVE_BASE = r"C:\Users\Trader\Documents\AJZStrategiesLLC"

# Tracker Dashboard location (for auto-matching System names)
TRACKER_DASHBOARD_PATH = r"C:\Users\Trader\AI-Agent-Learning-Hub\projects\P_115_BuytheDipTradingSystem\data\P_115_TrackerDashboard_V3.xlsx"

# Input/Output folders
PAPER_INPUT_FOLDER = os.path.join(AI_HUB_BASE, "data", "processed", "paper")
LIVE_INPUT_FOLDER = os.path.join(AI_HUB_BASE, "data", "processed", "live")
PAPER_OUTPUT_FOLDER = os.path.join(AI_HUB_BASE, "data")  # Paper logs in data/ folder
LIVE_OUTPUT_FOLDER = os.path.join(LIVE_BASE, "2026_Operations")

# Output Excel log files
EXCEL_LOGS = {
    "options_paper": os.path.join(PAPER_OUTPUT_FOLDER, "D_020_2026_AJZ_Strategies_Options_Log_V1.xlsx"),
    "stocks_paper": os.path.join(PAPER_OUTPUT_FOLDER, "D_020_2026_AJZ_Strategies_Stock_Log_V1.xlsx"),
    "options_live": os.path.join(LIVE_OUTPUT_FOLDER, "P_020_2026_AJZ_Strategies_Options_Log_v1.xlsx"),
    "stocks_live": os.path.join(LIVE_OUTPUT_FOLDER, "P_020_2026_AJZ_Strategies_Stock_Log_v1.xlsx")
}

# Sheet names in Excel logs
OPTIONS_SHEET_NAME = "Trade Log 2026"
STOCKS_SHEET_NAME = "Trade Log"

# Schema folder - drives all column mapping, no hardcoded column letters
SCHEMA_FOLDER = os.path.join(AI_HUB_BASE, "data", "schemas")

def load_schema(trade_type):
    filename = "options_schema.json" if trade_type == "options" else "stocks_schema.json"
    schema_path = os.path.join(SCHEMA_FOLDER, filename)
    try:
        with open(schema_path, 'r', encoding='utf-8-sig') as f:
            schema = json.load(f)
        print(f"  Loaded schema: {filename} v{schema.get(chr(118)+'ersion','?')} ({len(schema['columns'])} columns)")
        return schema['columns']
    except FileNotFoundError:
        print(f"  ERROR: Schema file not found: {schema_path}")
        return None
    except Exception as e:
        print(f"  ERROR loading schema: {e}")
        return None

def build_column_map_from_excel(ws, schema_columns):
    excel_headers = {}
    for cell in ws[1]:
        if cell.value:
            excel_headers[str(cell.value).strip()] = cell.column_letter
    column_map = {}
    formula_columns = []
    for col_def in schema_columns:
        name = col_def['name']
        source = col_def['source']
        if name in excel_headers:
            col_letter = excel_headers[name]
            if source == 'formula':
                formula_columns.append(col_letter)
            else:
                column_map[name] = col_letter
        elif source == 'parser':
            print(f"  Warning: Schema column '{name}' not found in Excel template - skipping")
    print(f"  Column map: {len(column_map)} data columns, {len(formula_columns)} formula columns protected")
    return column_map, formula_columns

# ============================================================
# TRACKER DASHBOARD MATCHING (REQ-020126_01)
# ============================================================

def load_tracker_dashboard():
    """
    Load Tracker Dashboard to match System names.
    Returns DataFrame with Symbol, Date, and Signal Source columns.
    """
    try:
        if not os.path.exists(TRACKER_DASHBOARD_PATH):
            print(f"  Ã¢Å¡Â  Warning: Tracker Dashboard not found at:")
            print(f"     {TRACKER_DASHBOARD_PATH}")
            print(f"  Ã¢â€ â€™ System names will default to 'TOS_Import'")
            return None
        
        df = pd.read_excel(TRACKER_DASHBOARD_PATH, sheet_name="Tracker Log")
        
        # Find the columns we need (flexible column name matching)
        symbol_col = None
        date_col = None
        signal_col = None
        
        for col in df.columns:
            col_lower = str(col).lower()
            if ('symbol' in col_lower or 'buy' in col_lower) and 'date' not in col_lower:
                symbol_col = col
            elif 'date' in col_lower or 'buy date' in col_lower:
                date_col = col
            elif 'signal' in col_lower or 'source' in col_lower:
                signal_col = col
        
        if not all([symbol_col, date_col, signal_col]):
            print(f"  Ã¢Å¡Â  Warning: Could not find required columns in Tracker Dashboard")
            print(f"     Looking for: Symbol/Buy, Date, Signal Source")
            print(f"  Ã¢â€ â€™ System names will default to 'TOS_Import'")
            return None
        
        # Create simplified lookup DataFrame
        tracker = df[[symbol_col, date_col, signal_col]].copy()
        tracker.columns = ['Symbol', 'Date', 'System']
        
        # Clean up Symbol (remove spaces, uppercase)
        tracker['Symbol'] = tracker['Symbol'].astype(str).str.strip().str.upper()
        
        # Convert dates to datetime
        tracker['Date'] = pd.to_datetime(tracker['Date'], errors='coerce')
        
        # Remove rows with missing data
        tracker = tracker.dropna()
        
        print(f"  Ã¢Å“â€œ Loaded Tracker Dashboard: {len(tracker)} trades")
        return tracker
        
    except Exception as e:
        print(f"  Ã¢Å¡Â  Warning: Error loading Tracker Dashboard: {e}")
        print(f"  Ã¢â€ â€™ System names will default to 'TOS_Import'")
        return None


def match_system_name(symbol, trade_date, tracker_df, trade_type="stocks"):
    """
    Match Symbol + Date against Tracker Dashboard to get System name.
    
    Args:
        symbol: Stock/option symbol (e.g., "QBTS", "AAPL")
        trade_date: Trade date (string or datetime)
        tracker_df: Tracker Dashboard DataFrame
    
    Returns:
        System name (e.g., "P_115", "P_118") or "TOS_Import" if no match
    """
    if tracker_df is None or tracker_df.empty:
        return "TOS_Import"
    
    try:
        # Clean symbol (remove option details, just keep base symbol)
        clean_symbol = re.sub(r'\s+\d+.*', '', str(symbol)).strip().upper()
        
        # Convert trade_date to datetime
        if isinstance(trade_date, str):
            trade_date = pd.to_datetime(trade_date, errors='coerce')
        
        if pd.isna(trade_date):
            return "TOS_Import"
        
        if trade_type == 'options':
            # Options: symbol only, signal date within 3 days of trade date
            symbol_matches = tracker_df[tracker_df['Symbol'].str.upper() == clean_symbol].copy()
            if symbol_matches.empty:
                return 'TOS_Import'
            symbol_matches['days_diff'] = (symbol_matches['Date'] - trade_date).dt.days.abs()
            matches = symbol_matches[symbol_matches['days_diff'] <= 3].sort_values('days_diff')
        else:
            # Stocks: symbol + 3-day window, sorted by Symbol then Date
            symbol_matches = tracker_df[tracker_df['Symbol'].str.upper() == clean_symbol].copy()
            if symbol_matches.empty:
                return 'TOS_Import'
            symbol_matches = symbol_matches.sort_values(['Symbol', 'Date'])
            symbol_matches['days_diff'] = (symbol_matches['Date'] - trade_date).dt.days.abs()
            matches = symbol_matches[symbol_matches['days_diff'] <= 3].sort_values('days_diff')
        if len(matches) > 0:
            # Use first match if multiple found
            return matches.iloc[0]['System']
        else:
            return "TOS_Import"
            
    except Exception as e:
        print(f"  Ã¢Å¡Â  Error matching {symbol} on {trade_date}: {e}")
        return "TOS_Import"


# ============================================================
# CSV AND EXCEL FUNCTIONS
# ============================================================

def find_latest_csv(folder, pattern):
    """
    Find most recent CSV file matching pattern in folder.
    
    Args:
        folder: Folder path to search
        pattern: Pattern to match (e.g., "*OPTIONS_IMPORT.csv")
    
    Returns:
        Full path to most recent file, or None if not found
    """
    import glob
    
    search_pattern = os.path.join(folder, pattern)
    files = glob.glob(search_pattern)
    
    if not files:
        return None
    
    # Sort by modification time, newest first
    files.sort(key=os.path.getmtime, reverse=True)
    return files[0]


def read_csv_data(csv_path):
    """
    Read CSV file and return as DataFrame.
    Handles various CSV formats and encodings.
    """
    try:
        for encoding in ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
            try:
                df = pd.read_csv(csv_path, encoding=encoding)
                print(f"  Ã¢Å“â€œ Read CSV: {os.path.basename(csv_path)} ({len(df)} rows)")
                return df
            except UnicodeDecodeError:
                continue
        
        raise Exception("Could not read CSV with any known encoding")
    
    except Exception as e:
        print(f"  Ã¢Å“â€” Error reading CSV: {e}")
        return None


def clear_data_columns(ws, columns_to_clear, start_row=2, end_row=101):
    """
    Clear data in specified columns (preserve formula columns).
    
    Args:
        ws: openpyxl worksheet
        columns_to_clear: List of column letters to clear (e.g., ['A', 'B', 'C'])
        start_row: First row to clear (default 2, after header)
        end_row: Last row to clear (default 101)
    """
    for row in range(start_row, end_row + 1):
        for col in columns_to_clear:
            ws[f"{col}{row}"] = None


def import_trades(csv_path, excel_path, trade_type="options", tracker_df=None):
    """
    Import trades from CSV to Excel log workbook.
    Preserves formulas and auto-matches System names from Tracker.
    
    Args:
        csv_path: Path to parser output CSV
        excel_path: Path to Excel log file
        trade_type: "options" or "stocks"
        tracker_df: Tracker Dashboard DataFrame for auto-matching
    
    Returns:
        True if successful, False otherwise
    """
    print(f"\n{'='*60}")
    print(f"IMPORTING {trade_type.upper()} TRADES")
    print(f"{'='*60}")
    print(f"Input:  {os.path.basename(csv_path)}")
    print(f"Output: {os.path.basename(excel_path)}")
    
    # Check files exist
    if not os.path.exists(csv_path):
        print(f"  Ã¢Å“â€” ERROR: CSV file not found: {csv_path}")
        return False
    
    if not os.path.exists(excel_path):
        print(f"  Ã¢Å“â€” ERROR: Excel file not found: {excel_path}")
        return False
    
    # Read CSV
    print(f"\n  Reading CSV data...")
    df = read_csv_data(csv_path)
    
    if df is None or df.empty:
        print(f"  Ã¢Å“â€” ERROR: No data in CSV")
        return False
    
    # Auto-match System names from Tracker Dashboard
    if tracker_df is not None and 'System' in df.columns:
        print(f"\n  Auto-matching System names from Tracker Dashboard...")
        matched = 0
        for idx, row in df.iterrows():
            if row['System'] == 'TOS_Import':
                symbol = row.get('Symbol', '')
                trade_date = row.get('Trade Date', '')
                matched_system = match_system_name(symbol, trade_date, tracker_df, trade_type)
                df.at[idx, 'System'] = matched_system
                if matched_system != 'TOS_Import':
                    matched += 1
        
        print(f"  Ã¢Å“â€œ Matched {matched}/{len(df)} trades to specific systems")
        if matched < len(df):
            print(f"  Ã¢â€ â€™ {len(df) - matched} trades remain as 'TOS_Import' (no Tracker match)")
    
    # Determine sheet name based on trade type
    if trade_type == "options":
        sheet_name = OPTIONS_SHEET_NAME
    else:
        sheet_name = STOCKS_SHEET_NAME



    # Open Excel workbook
    print(f"\n  Opening Excel workbook...")
    try:
        wb = openpyxl.load_workbook(excel_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        print(f"  Ã¢Å“â€œ Opened sheet: {ws.title}")
    except Exception as e:
        print(f"  Ã¢Å“â€” ERROR: Could not open workbook: {e}")
        return False
    
    # Load schema and build column map dynamically from Excel headers
    schema_columns = load_schema(trade_type)
    if schema_columns is None:
        return False
    column_map, formula_columns = build_column_map_from_excel(ws, schema_columns)
    date_columns = [col['name'] for col in schema_columns if col.get('type') == 'date']

    # Clear existing data (preserve formula columns from schema)
    print(f"\n  Clearing existing data (preserving formulas)...")
    clear_data_columns(ws, list(column_map.values()), start_row=2, end_row=101)
    print("  Data cleared")

    # Write CSV data to Excel using schema-driven column mapping
    print("\n  Writing data to Excel...")

    rows_written = 0

    for idx, row in df.iterrows():
        excel_row = idx + 2
        if excel_row > 101:
            print("  Warning: More than 100 rows. Truncating at row 101.")
            break
        for csv_col, excel_col in column_map.items():
            if csv_col in df.columns:
                value = row[csv_col]
                if pd.isna(value):
                    value = None
                # Write value - convert dates to datetime for proper Excel formatting
                if value is not None and csv_col in date_columns:
                    try:
                        dt = pd.to_datetime(value, errors='coerce')
                        if not pd.isna(dt):
                            cell = ws[f"{excel_col}{excel_row}"]
                            cell.value = dt.to_pydatetime()
                            cell.number_format = 'DD-MMM-YY'
                        else:
                            ws[f"{excel_col}{excel_row}"] = value
                    except:
                        ws[f"{excel_col}{excel_row}"] = value
                else:
                    ws[f"{excel_col}{excel_row}"] = value
        rows_written += 1
    
    print(f"  Ã¢Å“â€œ Wrote {rows_written} rows")
    
    # Save workbook
    print(f"\n  Saving workbook...")
    try:
        wb.save(excel_path)
        print(f"  Ã¢Å“â€œ Saved successfully!")
    except Exception as e:
        print(f"  Ã¢Å“â€” ERROR: Could not save workbook: {e}")
        print(f"     (Is the file open in Excel? Close it and try again.)")
        return False
    
    wb.close()
    
    print(f"\n{'='*60}")
    print(f"Ã¢Å“â€¦ SUCCESS: {rows_written} trades imported to {os.path.basename(excel_path)}")
    print(f"{'='*60}")
    
    return True


# ============================================================
# MAIN MENU
# ============================================================

def main():
    """Main menu for trade import utility."""
    
    # Load Tracker Dashboard once at startup
    print("\n" + "="*60)
    print("   P_020 TRADE IMPORT UTILITY v2.2")
    print("="*60)
    print("\n  Loading Tracker Dashboard for auto-matching...")
    tracker_df = load_tracker_dashboard()
    
    while True:
        print("\n" + "="*60)
        print("   IMPORT MENU")
        print("="*60)
        print(f"\n   Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("\n   Select import type:\n")
        print("   [1] Options Ã¢â€ â€™ Paper Log")
        print("   [2] Options Ã¢â€ â€™ Live Log")
        print("   [3] Stocks  Ã¢â€ â€™ Paper Log")
        print("   [4] Stocks  Ã¢â€ â€™ Live Log")
        print("   [5] Import ALL Ã¢â€ â€™ Paper")
        print("   [6] Import ALL Ã¢â€ â€™ Live")
        print("   [7] Reload Tracker Dashboard")
        print("   [0] Exit")
        
        choice = input("\n   Enter choice (0-7): ").strip()
        
        if choice == "1":
            csv = find_latest_csv(PAPER_INPUT_FOLDER, "*OPTIONS_IMPORT.csv")
            if csv:
                import_trades(csv, EXCEL_LOGS["options_paper"], "options", tracker_df)
            else:
                print(f"\n  Ã¢Å“â€” No OPTIONS CSV found in {PAPER_INPUT_FOLDER}")
        
        elif choice == "2":
            csv = find_latest_csv(LIVE_INPUT_FOLDER, "*OPTIONS_IMPORT.csv")
            if csv:
                import_trades(csv, EXCEL_LOGS["options_live"], "options", tracker_df)
            else:
                print(f"\n  Ã¢Å“â€” No OPTIONS CSV found in {LIVE_INPUT_FOLDER}")
        
        elif choice == "3":
            csv = find_latest_csv(PAPER_INPUT_FOLDER, "*STOCKS_IMPORT.csv")
            if csv:
                import_trades(csv, EXCEL_LOGS["stocks_paper"], "stocks", tracker_df)
            else:
                print(f"\n  Ã¢Å“â€” No STOCKS CSV found in {PAPER_INPUT_FOLDER}")
        
        elif choice == "4":
            csv = find_latest_csv(LIVE_INPUT_FOLDER, "*STOCKS_IMPORT.csv")
            if csv:
                import_trades(csv, EXCEL_LOGS["stocks_live"], "stocks", tracker_df)
            else:
                print(f"\n  Ã¢Å“â€” No STOCKS CSV found in {LIVE_INPUT_FOLDER}")
        
        elif choice == "5":
            # Import all to paper
            csv_opt = find_latest_csv(PAPER_INPUT_FOLDER, "*OPTIONS_IMPORT.csv")
            csv_stk = find_latest_csv(PAPER_INPUT_FOLDER, "*STOCKS_IMPORT.csv")
            if csv_opt:
                import_trades(csv_opt, EXCEL_LOGS["options_paper"], "options", tracker_df)
            if csv_stk:
                import_trades(csv_stk, EXCEL_LOGS["stocks_paper"], "stocks", tracker_df)
        
        elif choice == "6":
            # Import all to live
            csv_opt = find_latest_csv(LIVE_INPUT_FOLDER, "*OPTIONS_IMPORT.csv")
            csv_stk = find_latest_csv(LIVE_INPUT_FOLDER, "*STOCKS_IMPORT.csv")
            if csv_opt:
                import_trades(csv_opt, EXCEL_LOGS["options_live"], "options", tracker_df)
            if csv_stk:
                import_trades(csv_stk, EXCEL_LOGS["stocks_live"], "stocks", tracker_df)
        
        elif choice == "7":
            print("\n  Reloading Tracker Dashboard...")
            tracker_df = load_tracker_dashboard()
        
        elif choice == "0":
            print("\n   Goodbye!")
            break
        
        else:
            print("\n   Ã¢Å¡Â  Invalid choice. Please try again.")
        
        input("\n   Press Enter to continue...")


if __name__ == "__main__":
    main()




