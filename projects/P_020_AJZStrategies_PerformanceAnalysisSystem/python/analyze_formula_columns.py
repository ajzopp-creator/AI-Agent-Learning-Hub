import openpyxl
from openpyxl.utils import get_column_letter

def analyze_formulas(file_path, sheet_name=None):
    """
    Analyze an Excel file to find all columns containing formulas.
    """
    print(f"\n{'='*70}")
    print(f"Analyzing: {file_path}")
    print(f"{'='*70}")
    
    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb.active if sheet_name is None else wb[sheet_name]
    
    # Check first 10 data rows for formulas
    formula_columns = set()
    
    for row_num in range(2, 12):  # Check rows 2-11 (first 10 data rows)
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                col_letter = get_column_letter(col_num)
                formula_columns.add(col_letter)
                print(f"  Row {row_num}, Col {col_letter}: {cell.value[:50]}...")
    
    print(f"\n{'='*70}")
    print(f"FORMULA COLUMNS FOUND: {sorted(formula_columns)}")
    print(f"{'='*70}\n")
    
    return sorted(formula_columns)

# Analyze current D_020 files (restored from backup)
options_file = r"C:\Users\Trader\AI-Agent-Learning-Hub\projects\P_020_AJZStrategies_PerformanceAnalysisSystem\data\D_020_2026_AJZ_Strategies_Options_Log_V1.xlsx"
print("\n### ANALYZING OPTIONS FILE ###")
options_formulas = analyze_formulas(options_file)

stocks_file = r"C:\Users\Trader\AI-Agent-Learning-Hub\projects\P_020_AJZStrategies_PerformanceAnalysisSystem\data\D_020_2026_AJZ_Strategies_Stock_Log_V1.xlsx"
print("\n### ANALYZING STOCKS FILE ###")
try:
    stocks_formulas = analyze_formulas(stocks_file)
except Exception as e:
    print(f"Could not analyze stocks file: {e}")
    stocks_formulas = []

print("\n" + "="*70)
print("SUMMARY - Formula Columns to Protect:")
print("="*70)
print(f"Options: {options_formulas}")
print(f"Stocks:  {stocks_formulas}")
print("="*70)
