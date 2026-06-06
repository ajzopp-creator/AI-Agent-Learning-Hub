"""
P_020 Trade Import Utility - Enhanced Version
==============================================
Imports trade data from TOS export CSV files to Excel log workbooks.

Features:
- Automatically matches System names from Tracker Dashboard (REQ-020126_01)
- Preserves formulas in calculated columns (REQ-020207_01)
- Handles both paper and live accounts
- Menu-driven interface

Requirements:
    pip install openpyxl pandas

Author: Anthony (AJZ Strategies LLC)
Version: 2.2
Date: 2026-02-07
"""

import pandas as pd
import openpyxl
import os
from datetime import datetime
import re

# ============================================================
# CONFIGURATION
# ============================================================

# Base paths
AI_HUB_BASE = r"C:\Users\Trader\AI-Agent-Learning-Hub\projects\P_020_AJZStrategies_PerformanceAnalysisSystem"
LIVE_BASE = r"C:\Users\Trader\Documents\AJZStrategiesLLC"

# Tracker Dashboard location (for auto-matching System names)
TRACKER_DASHBOARD_PATH = r"C:\Users\Trader\Documents\AJZStrategiesLLC\P_115_TrackerAudit\P_115_118_TrackerDashboard_V2.xlsx"

# Input/Output folders
PAPER_INPUT_FOLDER = os.path.join(AI_HUB_BASE, "data", "processed", "paper")
LIVE_INPUT_FOLDER = os.path.join(AI_HUB_BASE, "data", "processed", "live")
PAPER_OUTPUT_FOLDER = os.path.join(AI_HUB_BASE, "data")  # Paper logs in data/ folder
LIVE_OUTPUT_FOLDER = os.path.join(LIVE_BASE, "2026_Operations")

# Output Excel log files
EXCEL_LOGS = {
    "options_paper": os.path.join(PAPER_OUTPUT_FOLDER, "D_020_2026_AJZ_Strategies_Options_Log_V1.xlsx"),
    "stocks_paper": os.path.join(PAPER_OUTPUT_FOLDER, "D_020_2026__AJZ_Strategies_Stock_Log_V1.xlsx"),
    "options_live": os.path.join(LIVE_OUTPUT_FOLDER, "P_020_2026_AJZ_Strategies_Options_Log_v1.xlsx"),
    "stocks_live": os.path.join(LIVE_OUTPUT_FOLDER, "P_020_2026_AJZ_Strategies_Stock_Log_v1.xlsx")
}

# Sheet name in Excel logs
LOG_SHEET_NAME = "Trade_Log"

# Columns to write (data columns only - preserve formulas!)
# Options: 20 columns total
OPTIONS_COLUMNS_TO_WRITE = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'R', 'S', 'T']
# Stocks: 26 columns total
STOCKS_COLUMNS_TO_WRITE = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

# Formula columns to preserve (DON'T overwrite these)
OPTIONS_FORMULA_COLUMNS = ['Q']  # Gain/Loss for options
STOCKS_FORMULA_COLUMNS = ['Q', 'R']  # Gain/Loss, ROI for stocks

# ============================================================
# TRACKER DASHBOARD MATCHING (REQ-020126_01)
# ============================================================

def load_tracker_dashboard():
    """
    Load Tracker Dashboard to match System names.
    Returns DataFrame with Symbol, Date, and Signal Source columns.
    """
    try:
        if not os.path.exists(TRACKER_DASHBOARD_PATH):
            print(f"  ⚠ Warning: Tracker Dashboard not found at:")
            print(f"     {TRACKER_DASHBOARD_PATH}")
            print(f"  → System names will default to 'TOS_Import'")
            return None
        
        df = pd.read_excel(TRACKER_DASHBOARD_PATH)
        
        # Find the columns we need (flexible column name matching)
        symbol_col = None
        date_col = None
        signal_col = None
        
        for col in df.columns:
            col_lower = str(col).lower()
            if 'buy' in col_lower and 'date' not in col_lower:
                symbol_col = col
            elif 'date' in col_lower or 'buy date' in col_lower:
                date_col = col
            elif 'signal' in col_lower or 'source' in col_lower:
                signal_col = col
        
        if not all([symbol_col, date_col, signal_col]):
            print(f"  ⚠ Warning: Could not find required columns in Tracker Dashboard")
            print(f"     Looking for: Symbol/Buy, Date, Signal Source")
            print(f"  → System names will default to 'TOS_Import'")
            return None
        
        # Create simplified lookup DataFrame
        tracker = df[[symbol_col, date_col, signal_col]].copy()
        tracker.columns = ['Symbol', 'Date', 'System']
        
        # Clean up Symbol (remove spaces, uppercase)
        tracker['Symbol'] = tracker['Symbol'].astype(str).str.strip().str.upper()
        
        # Convert dates to datetime
        tracker['Date'] = pd.to_datetime(tracker['Date'], errors='coerce')
        
        # Remove rows with missing data
        tracker = tracker.dropna()
        
        print(f"  ✓ Loaded Tracker Dashboard: {len(tracker)} trades")
        return tracker
        
    except Exception as e:
        print(f"  ⚠ Warning: Error loading Tracker Dashboard: {e}")
        print(f"  → System names will default to 'TOS_Import'")
        return None


def match_system_name(symbol, trade_date, tracker_df):
    """
    Match Symbol + Date against Tracker Dashboard to get System name.
    
    Args:
        symbol: Stock/option symbol (e.g., "QBTS", "AAPL")
        trade_date: Trade date (string or datetime)
        tracker_df: Tracker Dashboard DataFrame
    
    Returns:
        System name (e.g., "P_115", "P_118") or "TOS_Import" if no match
    """
    if tracker_df is None or tracker_df.empty:
        return "TOS_Import"
    
    try:
        # Clean symbol (remove option details, just keep base symbol)
        clean_symbol = re.sub(r'\s+\d+.*', '', str(symbol)).strip().upper()
        
        # Convert trade_date to datetime
        if isinstance(trade_date, str):
            trade_date = pd.to_datetime(trade_date, errors='coerce')
        
        if pd.isna(trade_date):
            return "TOS_Import"
        
        # Match Symbol + Date (case-insensitive, date exact)
        matches = tracker_df[
            (tracker_df['Symbol'].str.upper() == clean_symbol) &
            (tracker_df['Date'].dt.date == trade_date.date())
        ]
        
        if len(matches) > 0:
            # Use first match if multiple found
            return matches.iloc[0]['System']
        else:
            return "TOS_Import"
            
    except Exception as e:
        print(f"  ⚠ Error matching {symbol} on {trade_date}: {e}")
        return "TOS_Import"


# ============================================================
# CSV AND EXCEL FUNCTIONS
# ============================================================

def find_latest_csv(folder, pattern):
    """
    Find most recent CSV file matching pattern in folder.
    
    Args:
        folder: Folder path to search
        pattern: Pattern to match (e.g., "*OPTIONS_IMPORT.csv")
    
    Returns:
        Full path to most recent file, or None if not found
    """
    import glob
    
    search_pattern = os.path.join(folder, pattern)
    files = glob.glob(search_pattern)
    
    if not files:
        return None
    
    # Sort by modification time, newest first
    files.sort(key=os.path.getmtime, reverse=True)
    return files[0]


def read_csv_data(csv_path):
    """
    Read CSV file and return as DataFrame.
    Handles various CSV formats and encodings.
    """
    try:
        for encoding in ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
            try:
                df = pd.read_csv(csv_path, encoding=encoding)
                print(f"  ✓ Read CSV: {os.path.basename(csv_path)} ({len(df)} rows)")
                return df
            except UnicodeDecodeError:
                continue
        
        raise Exception("Could not read CSV with any known encoding")
    
    except Exception as e:
        print(f"  ✗ Error reading CSV: {e}")
        return None


def clear_data_columns(ws, columns_to_clear, start_row=2, end_row=101):
    """
    Clear data in specified columns (preserve formula columns).
    
    Args:
        ws: openpyxl worksheet
        columns_to_clear: List of column letters to clear (e.g., ['A', 'B', 'C'])
        start_row: First row to clear (default 2, after header)
        end_row: Last row to clear (default 101)
    """
    for row in range(start_row, end_row + 1):
        for col in columns_to_clear:
            ws[f"{col}{row}"] = None


def import_trades(csv_path, excel_path, trade_type="options", tracker_df=None):
    """
    Import trades from CSV to Excel log workbook.
    Preserves formulas and auto-matches System names from Tracker.
    
    Args:
        csv_path: Path to parser output CSV
        excel_path: Path to Excel log file
        trade_type: "options" or "stocks"
        tracker_df: Tracker Dashboard DataFrame for auto-matching
    
    Returns:
        True if successful, False otherwise
    """
    print(f"\n{'='*60}")
    print(f"IMPORTING {trade_type.upper()} TRADES")
    print(f"{'='*60}")
    print(f"Input:  {os.path.basename(csv_path)}")
    print(f"Output: {os.path.basename(excel_path)}")
    
    # Check files exist
    if not os.path.exists(csv_path):
        print(f"  ✗ ERROR: CSV file not found: {csv_path}")
        return False
    
    if not os.path.exists(excel_path):
        print(f"  ✗ ERROR: Excel file not found: {excel_path}")
        return False
    
    # Read CSV
    print(f"\n  Reading CSV data...")
    df = read_csv_data(csv_path)
    
    if df is None or df.empty:
        print(f"  ✗ ERROR: No data in CSV")
        return False
    
    # Auto-match System names from Tracker Dashboard
    if tracker_df is not None and 'System' in df.columns:
        print(f"\n  Auto-matching System names from Tracker Dashboard...")
        matched = 0
        for idx, row in df.iterrows():
            if row['System'] == 'TOS_Import':
                symbol = row.get('Symbol', '')
                trade_date = row.get('Trade Date', '')
                matched_system = match_system_name(symbol, trade_date, tracker_df)
                df.at[idx, 'System'] = matched_system
                if matched_system != 'TOS_Import':
                    matched += 1
        
        print(f"  ✓ Matched {matched}/{len(df)} trades to specific systems")
        if matched < len(df):
            print(f"  → {len(df) - matched} trades remain as 'TOS_Import' (no Tracker match)")
    
    # Open Excel workbook
    print(f"\n  Opening Excel workbook...")
    try:
        wb = openpyxl.load_workbook(excel_path)
        if LOG_SHEET_NAME in wb.sheetnames:
            ws = wb[LOG_SHEET_NAME]
        else:
            ws = wb.active
        print(f"  ✓ Opened sheet: {ws.title}")
    except Exception as e:
        print(f"  ✗ ERROR: Could not open workbook: {e}")
        return False
    
    # Determine which columns to write (based on trade type)
    if trade_type == "options":
        data_columns = [col for col in OPTIONS_COLUMNS_TO_WRITE if col not in OPTIONS_FORMULA_COLUMNS]
    else:  # stocks
        data_columns = [col for col in STOCKS_COLUMNS_TO_WRITE if col not in STOCKS_FORMULA_COLUMNS]
    
    # Clear existing data (preserve formulas)
    print(f"\n  Clearing existing data (preserving formulas)...")
    clear_data_columns(ws, data_columns, start_row=2, end_row=101)
    print(f"  ✓ Data cleared")
    
    # Write CSV data to Excel
    print(f"\n  Writing data to Excel...")
    rows_written = 0
    
    # Get CSV column names
    csv_columns = df.columns.tolist()
    
    for idx, row in df.iterrows():
        excel_row = idx + 2  # Start at row 2 (row 1 is header)
        
        if excel_row > 101:
            print(f"  ⚠ Warning: More than 100 rows. Truncating at row 101.")
            break
        
        # Map CSV columns to Excel columns
        for i, csv_col in enumerate(csv_columns):
            if i < len(data_columns):
                excel_col = data_columns[i]
                value = row[csv_col]
                
                # Handle NaN values
                if pd.isna(value):
                    value = None
                
                ws[f"{excel_col}{excel_row}"] = value
        
        rows_written += 1
    
    print(f"  ✓ Wrote {rows_written} rows")
    
    # Save workbook
    print(f"\n  Saving workbook...")
    try:
        wb.save(excel_path)
        print(f"  ✓ Saved successfully!")
    except Exception as e:
        print(f"  ✗ ERROR: Could not save workbook: {e}")
        print(f"     (Is the file open in Excel? Close it and try again.)")
        return False
    
    wb.close()
    
    print(f"\n{'='*60}")
    print(f"✅ SUCCESS: {rows_written} trades imported to {os.path.basename(excel_path)}")
    print(f"{'='*60}")
    
    return True


# ============================================================
# MAIN MENU
# ============================================================

def main():
    """Main menu for trade import utility."""
    
    # Load Tracker Dashboard once at startup
    print("\n" + "="*60)
    print("   P_020 TRADE IMPORT UTILITY v2.2")
    print("="*60)
    print("\n  Loading Tracker Dashboard for auto-matching...")
    tracker_df = load_tracker_dashboard()
    
    while True:
        print("\n" + "="*60)
        print("   IMPORT MENU")
        print("="*60)
        print(f"\n   Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("\n   Select import type:\n")
        print("   [1] Options → Paper Log")
        print("   [2] Options → Live Log")
        print("   [3] Stocks  → Paper Log")
        print("   [4] Stocks  → Live Log")
        print("   [5] Import ALL → Paper")
        print("   [6] Import ALL → Live")
        print("   [7] Reload Tracker Dashboard")
        print("   [0] Exit")
        
        choice = input("\n   Enter choice (0-7): ").strip()
        
        if choice == "1":
            csv = find_latest_csv(PAPER_INPUT_FOLDER, "*OPTIONS_IMPORT.csv")
            if csv:
                import_trades(csv, EXCEL_LOGS["options_paper"], "options", tracker_df)
            else:
                print(f"\n  ✗ No OPTIONS CSV found in {PAPER_INPUT_FOLDER}")
        
        elif choice == "2":
            csv = find_latest_csv(LIVE_INPUT_FOLDER, "*OPTIONS_IMPORT.csv")
            if csv:
                import_trades(csv, EXCEL_LOGS["options_live"], "options", tracker_df)
            else:
                print(f"\n  ✗ No OPTIONS CSV found in {LIVE_INPUT_FOLDER}")
        
        elif choice == "3":
            csv = find_latest_csv(PAPER_INPUT_FOLDER, "*STOCKS_IMPORT.csv")
            if csv:
                import_trades(csv, EXCEL_LOGS["stocks_paper"], "stocks", tracker_df)
            else:
                print(f"\n  ✗ No STOCKS CSV found in {PAPER_INPUT_FOLDER}")
        
        elif choice == "4":
            csv = find_latest_csv(LIVE_INPUT_FOLDER, "*STOCKS_IMPORT.csv")
            if csv:
                import_trades(csv, EXCEL_LOGS["stocks_live"], "stocks", tracker_df)
            else:
                print(f"\n  ✗ No STOCKS CSV found in {LIVE_INPUT_FOLDER}")
        
        elif choice == "5":
            # Import all to paper
            csv_opt = find_latest_csv(PAPER_INPUT_FOLDER, "*OPTIONS_IMPORT.csv")
            csv_stk = find_latest_csv(PAPER_INPUT_FOLDER, "*STOCKS_IMPORT.csv")
            if csv_opt:
                import_trades(csv_opt, EXCEL_LOGS["options_paper"], "options", tracker_df)
            if csv_stk:
                import_trades(csv_stk, EXCEL_LOGS["stocks_paper"], "stocks", tracker_df)
        
        elif choice == "6":
            # Import all to live
            csv_opt = find_latest_csv(LIVE_INPUT_FOLDER, "*OPTIONS_IMPORT.csv")
            csv_stk = find_latest_csv(LIVE_INPUT_FOLDER, "*STOCKS_IMPORT.csv")
            if csv_opt:
                import_trades(csv_opt, EXCEL_LOGS["options_live"], "options", tracker_df)
            if csv_stk:
                import_trades(csv_stk, EXCEL_LOGS["stocks_live"], "stocks", tracker_df)
        
        elif choice == "7":
            print("\n  Reloading Tracker Dashboard...")
            tracker_df = load_tracker_dashboard()
        
        elif choice == "0":
            print("\n   Goodbye!")
            break
        
        else:
            print("\n   ⚠ Invalid choice. Please try again.")
        
        input("\n   Press Enter to continue...")


if __name__ == "__main__":
    main()
